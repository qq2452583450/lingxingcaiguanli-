import json
import builtins
from io import BytesIO

from openpyxl import load_workbook


def create_inquiry_delete_tables(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_inquiry_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            inquiry_id INTEGER,
            material_id INTEGER,
            quantity REAL DEFAULT 1,
            tax_rate REAL DEFAULT 0.01,
            is_national_standard INTEGER DEFAULT 0,
            is_cash_price INTEGER DEFAULT 0,
            detail_spec TEXT,
            brand TEXT,
            create_time TEXT
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_inquiry_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER,
            supplier_id INTEGER,
            tax_price REAL DEFAULT 0,
            tax_exempt_price REAL DEFAULT 0,
            tax_rate REAL DEFAULT 0.13,
            total_amount REAL DEFAULT 0,
            is_lowest INTEGER DEFAULT 0,
            is_selected INTEGER DEFAULT 0,
            create_time TEXT
        )
        """
    )


def ensure_project_id_column(cursor):
    columns = [row[1] for row in cursor.execute("PRAGMA table_info(purchase_inquiries)").fetchall()]
    if "project_id" not in columns:
        cursor.execute("ALTER TABLE purchase_inquiries ADD COLUMN project_id INTEGER")


def ensure_stock_in_project_id_column(cursor):
    columns = [row[1] for row in cursor.execute("PRAGMA table_info(stock_in_orders)").fetchall()]
    if "project_id" not in columns:
        cursor.execute("ALTER TABLE stock_in_orders ADD COLUMN project_id INTEGER")


def seed_role_user(cursor, role_name, username, real_name):
    cursor.execute("SELECT id FROM roles WHERE role_name = ?", (role_name,))
    row = cursor.fetchone()
    if row:
        role_id = row[0]
    else:
        cursor.execute("INSERT INTO roles (role_name, permissions) VALUES (?, ?)", (role_name, ""))
        role_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (username, "x", real_name, role_id, 1, "2026-01-01 09:00:00"),
    )
    return cursor.lastrowid


def seed_special_approval_users(cursor):
    admin_id = seed_role_user(cursor, "系统管理员", "admin", "管理员")
    lei_id = seed_role_user(cursor, "材料审批负责人", "leikefeng", "雷克峰")
    tan_id = seed_role_user(cursor, "材料审批负责人", "tanxiang", "谭香")
    wang_id = seed_role_user(cursor, "材料员", "wanglihua", "王利华")
    return admin_id, lei_id, tan_id, wang_id


def seed_inquiry(cursor, inquiry_no, applicant_id, status="待审批", project_id=None):
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, project_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (inquiry_no, "2026-06-06", applicant_id, project_id, 0, 0, status, "2026-06-06 10:00:00"),
    )
    return cursor.lastrowid


def seed_project(cursor, project_code, project_name):
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        (project_code, project_name, "2026-01-01 09:00:00"),
    )
    return cursor.lastrowid


def set_session_user(client, user_id, username, real_name, role_name):
    with client.session_transaction() as sess:
        sess["user"] = {
            "id": user_id,
            "username": username,
            "real_name": real_name,
            "role_name": role_name,
        }


def test_purchase_inquiries_support_list_filters(client, test_db):
    cursor = test_db.cursor()
    cursor.execute("INSERT INTO roles (role_name, permissions) VALUES (?, ?)", ("系统管理员", ""))
    role_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("admin", "x", "管理员", role_id, 1, "2026-01-01 09:00:00"),
    )
    admin_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("buyer", "x", "张三", role_id, 1, "2026-01-01 09:00:00"),
    )
    buyer_id = cursor.lastrowid
    rows = [
        ("XJ-202606-001", "2026-06-01", buyer_id, 1200, 1, "待审批", "2026-06-01 10:00:00"),
        ("XJ-202606-002", "2026-06-03", admin_id, 900, 0, "草稿", "2026-06-03 10:00:00"),
        ("XJ-202605-003", "2026-05-28", buyer_id, 500, 1, "已同意", "2026-05-28 10:00:00"),
    ]
    cursor.executemany(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    test_db.commit()

    with client.session_transaction() as sess:
        sess["user"] = {"id": admin_id, "username": "admin", "real_name": "管理员"}

    response = client.get(
        "/api/purchase-inquiries?"
        "keyword=202606&applicant=张三&status=待审批&"
        "start_date=2026-06-01&end_date=2026-06-30&is_below=1"
    )

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is True
    assert [row["inquiry_no"] for row in data["data"]] == ["XJ-202606-001"]


def test_submit_draft_regenerates_generic_inquiry_no_for_final_project(client, test_db):
    cursor = test_db.cursor()
    clerk_id = seed_role_user(cursor, "材料员", "zhengrongjie", "郑荣杰")
    project_id = seed_project(cursor, "KMJJYC", "京江隐翠")
    cursor.execute("INSERT INTO units (unit_name, unit_code) VALUES (?, ?)", ("个", "GE"))
    unit_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, unit_id, tax_price, create_time)
        VALUES (?, ?, ?, ?, ?)
        """,
        ("KMLX00001", "测试材料", unit_id, 10, "2026-06-11 10:00:00"),
    )
    material_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO suppliers (supplier_name, create_time) VALUES (?, ?)",
        ("测试供应商", "2026-06-11 10:00:00"),
    )
    supplier_id = cursor.lastrowid
    draft_id = seed_inquiry(cursor, "CGXJ-260612-001", clerk_id, status="草稿", project_id=None)
    test_db.commit()
    set_session_user(client, clerk_id, "zhengrongjie", "郑荣杰", "材料员")

    response = client.post(
        f"/api/purchase-inquiries/draft/{draft_id}/submit",
        json={
            "project_id": project_id,
            "inquiry_date": "2026-06-11",
            "items": [
                {
                    "material_id": material_id,
                    "quantity": 1,
                    "library_price": 10,
                    "quotes": [
                        {
                            "supplier_id": supplier_id,
                            "tax_price": 12,
                            "tax_rate": 0.13,
                            "is_lowest": 1,
                        }
                    ],
                }
            ],
        },
    )

    data = response.get_json()
    assert data["success"] is True
    assert data["inquiry_no"] == "KMJJYC-20260611-001"
    row = test_db.execute("SELECT inquiry_no FROM purchase_inquiries WHERE id = ?", (draft_id,)).fetchone()
    assert row["inquiry_no"] == "KMJJYC-20260611-001"


def test_submit_draft_persists_selected_quote_and_total_amount(client, test_db):
    cursor = test_db.cursor()
    clerk_id = seed_role_user(cursor, "材料员", "quote_clerk", "材料员")
    project_id = seed_project(cursor, "CD-QUOTE", "成都询价项目")
    cursor.execute("INSERT INTO units (unit_name, unit_code) VALUES (?, ?)", ("个", "GE"))
    unit_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, unit_id, tax_price, create_time)
        VALUES (?, ?, ?, ?, ?)
        """,
        ("CDLX00011", "报价材料", unit_id, 20, "2026-06-11 10:00:00"),
    )
    material_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name, create_time) VALUES (?, ?)", ("高价供应商", "2026-06-11 10:00:00"))
    high_supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name, create_time) VALUES (?, ?)", ("低价供应商", "2026-06-11 10:00:00"))
    low_supplier_id = cursor.lastrowid
    draft_id = seed_inquiry(cursor, "XJ-DRAFT-SELECTED", clerk_id, status="草稿", project_id=project_id)
    test_db.commit()
    set_session_user(client, clerk_id, "quote_clerk", "材料员", "材料员")

    response = client.post(
        f"/api/purchase-inquiries/draft/{draft_id}/submit",
        json={
            "project_id": project_id,
            "inquiry_date": "2026-06-11",
            "items": [
                {
                    "material_id": material_id,
                    "quantity": 3,
                    "library_price": 20,
                    "selected_quote_id": low_supplier_id,
                    "quotes": [
                        {"supplier_id": high_supplier_id, "tax_price": 18, "tax_rate": 0.13},
                        {"supplier_id": low_supplier_id, "tax_price": 12, "tax_rate": 0.13},
                    ],
                }
            ],
        },
    )

    data = response.get_json()
    assert data["success"] is True
    inquiry = test_db.execute("SELECT total_amount FROM purchase_inquiries WHERE id = ?", (draft_id,)).fetchone()
    assert inquiry["total_amount"] == 36
    selected = test_db.execute(
        """
        SELECT q.supplier_id, q.is_selected
        FROM purchase_inquiry_quotes q
        JOIN purchase_inquiry_items i ON i.id = q.item_id
        WHERE i.inquiry_id = ? AND q.supplier_id = ?
        """,
        (draft_id, low_supplier_id),
    ).fetchone()
    assert selected["is_selected"] == 1


def test_purchase_inquiries_include_project_city_code_and_name(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    admin_id = seed_role_user(cursor, "系统管理员", "admin", "管理员")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-QUOTE", "广西报价项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-GX-PROJECT", admin_id, project_id=project_id)
    test_db.commit()
    set_session_user(client, admin_id, "admin", "管理员", "系统管理员")

    list_response = client.get("/api/purchase-inquiries")
    list_data = list_response.get_json()
    listed = next(row for row in list_data["data"] if row["id"] == inquiry_id)
    assert listed["project_city"] == "广西"
    assert listed["project_code"] == "GX-QUOTE"
    assert listed["project_name"] == "广西报价项目"

    detail_response = client.get(f"/api/purchase-inquiries/{inquiry_id}")
    detail_data = detail_response.get_json()
    assert detail_data["data"]["project_city"] == "广西"
    assert detail_data["data"]["project_code"] == "GX-QUOTE"
    assert detail_data["data"]["project_name"] == "广西报价项目"


def test_supplier_can_see_inquiries_where_selected_without_project_binding(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    clerk_id = seed_role_user(cursor, "材料员", "clerk", "材料员")
    supplier_user_id = seed_role_user(cursor, "供应商", "supplier_00001", "报价供应商")
    other_supplier_user_id = seed_role_user(cursor, "供应商", "supplier_00002", "其他供应商")
    cursor.execute(
        "INSERT INTO suppliers (supplier_name, user_id, create_time) VALUES (?, ?, ?)",
        ("报价供应商", supplier_user_id, "2026-01-01 09:00:00"),
    )
    supplier_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO suppliers (supplier_name, user_id, create_time) VALUES (?, ?, ?)",
        ("其他供应商", other_supplier_user_id, "2026-01-01 09:00:00"),
    )
    other_supplier_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-SUPPLIER", "广西供应商报价项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name, unit_code) VALUES (?, ?)", ("个", "PCS"))
    unit_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO materials (material_code, material_name, unit_id, create_time) VALUES (?, ?, ?, ?)",
        ("GXLX00001", "测试材料", unit_id, "2026-01-01 09:00:00"),
    )
    material_id = cursor.lastrowid
    visible_inquiry_id = seed_inquiry(cursor, "XJ-SUPPLIER-VISIBLE", clerk_id, project_id=project_id)
    hidden_inquiry_id = seed_inquiry(cursor, "XJ-SUPPLIER-HIDDEN", clerk_id, project_id=project_id)
    cursor.execute(
        "INSERT INTO purchase_inquiry_items (inquiry_id, material_id, quantity, create_time) VALUES (?, ?, ?, ?)",
        (visible_inquiry_id, material_id, 3, "2026-06-06 10:00:00"),
    )
    visible_item_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO purchase_inquiry_quotes (item_id, supplier_id, tax_price, total_amount, create_time) VALUES (?, ?, ?, ?, ?)",
        (visible_item_id, supplier_id, 12, 36, "2026-06-06 10:00:00"),
    )
    cursor.execute(
        "INSERT INTO purchase_inquiry_quotes (item_id, supplier_id, tax_price, total_amount, create_time) VALUES (?, ?, ?, ?, ?)",
        (visible_item_id, other_supplier_id, 14, 42, "2026-06-06 10:00:00"),
    )
    cursor.execute(
        "INSERT INTO purchase_inquiry_items (inquiry_id, material_id, quantity, create_time) VALUES (?, ?, ?, ?)",
        (hidden_inquiry_id, material_id, 3, "2026-06-06 10:00:00"),
    )
    hidden_item_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO purchase_inquiry_quotes (item_id, supplier_id, tax_price, total_amount, create_time) VALUES (?, ?, ?, ?, ?)",
        (hidden_item_id, other_supplier_id, 14, 42, "2026-06-06 10:00:00"),
    )
    test_db.commit()
    set_session_user(client, supplier_user_id, "supplier_00001", "报价供应商", "供应商")

    list_response = client.get("/api/purchase-inquiries")
    list_data = list_response.get_json()
    assert list_response.status_code == 200
    assert [row["inquiry_no"] for row in list_data["data"]] == ["XJ-SUPPLIER-VISIBLE"]
    assert list_data["data"][0]["project_display_name"] == "广西 / GX-SUPPLIER / 广西供应商报价项目"

    detail_response = client.get(f"/api/purchase-inquiries/{visible_inquiry_id}")
    assert detail_response.status_code == 200
    detail_data = detail_response.get_json()
    assert detail_data["data"]["inquiry_no"] == "XJ-SUPPLIER-VISIBLE"
    assert [quote["supplier_id"] for quote in detail_data["items"][0]["quotes"]] == [supplier_id]

    forbidden_response = client.get(f"/api/purchase-inquiries/{hidden_inquiry_id}")
    assert forbidden_response.status_code == 403


def test_material_clerk_can_delete_draft_inquiry(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    cursor.execute("INSERT INTO roles (role_name, permissions) VALUES (?, ?)", ("材料员", ""))
    role_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("clerk", "x", "材料员", role_id, 1, "2026-01-01 09:00:00"),
    )
    clerk_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("XJ-DRAFT-001", "2026-06-06", clerk_id, 0, 0, "草稿", "2026-06-06 10:00:00"),
    )
    inquiry_id = cursor.lastrowid
    test_db.commit()

    with client.session_transaction() as sess:
        sess["user"] = {
            "id": clerk_id,
            "username": "clerk",
            "real_name": "材料员",
            "role_name": "材料员",
        }

    response = client.delete(f"/api/purchase-inquiries/{inquiry_id}")

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is True
    cursor.execute("SELECT COUNT(*) FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == 0


def test_material_clerk_cannot_delete_non_draft_inquiry(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    cursor.execute("INSERT INTO roles (role_name, permissions) VALUES (?, ?)", ("材料员", ""))
    role_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("clerk", "x", "材料员", role_id, 1, "2026-01-01 09:00:00"),
    )
    clerk_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("XJ-PENDING-001", "2026-06-06", clerk_id, 0, 0, "待审批", "2026-06-06 10:00:00"),
    )
    inquiry_id = cursor.lastrowid
    test_db.commit()

    with client.session_transaction() as sess:
        sess["user"] = {
            "id": clerk_id,
            "username": "clerk",
            "real_name": "材料员",
            "role_name": "材料员",
        }

    response = client.delete(f"/api/purchase-inquiries/{inquiry_id}")

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is False
    cursor.execute("SELECT COUNT(*) FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == 1


def test_approval_print_uses_project_management_and_project_manager_labels(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    try:
        cursor.execute("ALTER TABLE purchase_inquiries ADD COLUMN project_id INTEGER")
    except Exception:
        pass  # 列已存在
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("buyer", "x", "申请人", 1, 1, "2026-01-01 09:00:00"),
    )
    buyer_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("XJ-PRINT-001", "2026-06-06", buyer_id, 0, 0, "待审批", "2026-06-06 10:00:00"),
    )
    inquiry_id = cursor.lastrowid
    test_db.commit()

    response = client.get(f"/api/purchase-inquiries/{inquiry_id}/approval-print")

    assert response.status_code == 200
    html = response.data.decode("utf-8")
    assert "项目管理处" in html
    assert "项目负责人签字" in html
    assert "部门主管" not in html
    assert "主管签字" not in html
    assert "总经理签字" not in html


def test_approval_print_uses_excel_like_supplier_columns_and_selected_totals(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    applicant_id = seed_role_user(cursor, "材料员", "buyer_print", "郑荣杰")
    approver_id = seed_role_user(cursor, "材料审批负责人", "pm_approver", "雷克峰")
    project_id = seed_project(cursor, "KMJJYC", "京江隐翠")
    cursor.execute("INSERT INTO units (unit_name, unit_code) VALUES (?, ?)", ("个", "GE"))
    unit_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("佩文筛网",))
    supplier_a = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("捷阳五金",))
    supplier_b = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, specification, detail_spec, brand, unit_id, tax_price)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("KMLX00021", "钢丝网", "DN25", "加厚", "国标", unit_id, 13),
    )
    material_a = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, specification, detail_spec, brand, unit_id, tax_price)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("KMLX00022", "膨胀螺丝", "M8", "常规", "无", unit_id, 9),
    )
    material_b = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, project_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        ("CGXJ-260611-001", "2026-06-11", applicant_id, project_id, 61, 1, "已同意", "2026-06-11 10:00:00"),
    )
    inquiry_id = cursor.lastrowid
    for material_id, quantity, selected_supplier in [
        (material_a, 2, supplier_b),
        (material_b, 5, supplier_a),
    ]:
        cursor.execute(
            """
            INSERT INTO purchase_inquiry_items (
                inquiry_id, material_id, quantity, library_price, selected_quote_id,
                tax_rate, is_national_standard, detail_spec, brand, create_time
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (inquiry_id, material_id, quantity, 10, selected_supplier, 0.01, 1, "", "", "2026-06-11 10:00:00"),
        )
        item_id = cursor.lastrowid
        if material_id == material_a:
            quotes = [
                (supplier_a, 12, 24, 0, 0),
                (supplier_b, 10.5, 21, 1, 1),
            ]
        else:
            quotes = [
                (supplier_a, 8, 40, 1, 1),
                (supplier_b, 8.5, 42.5, 0, 0),
            ]
        for supplier_id, tax_price, total_amount, is_lowest, is_selected in quotes:
            cursor.execute(
                """
                INSERT INTO purchase_inquiry_quotes (
                    item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
                    total_amount, is_lowest, is_selected, create_time
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (item_id, supplier_id, tax_price, round(tax_price / 1.01, 2), 0.01, total_amount, is_lowest, is_selected, "2026-06-11 10:00:00"),
            )
    cursor.execute(
        """
        INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("purchase_inquiry", inquiry_id, approver_id, "雷克峰", "主管同意", "同意采购", "2026-06-11 15:20:00"),
    )
    test_db.commit()

    response = client.get(f"/api/purchase-inquiries/{inquiry_id}/approval-print")

    assert response.status_code == 200
    html = response.data.decode("utf-8")
    assert "零星材采购比价审批签字单" in html
    assert "项目名称：昆明 / KMJJYC / 京江隐翠" in html
    assert "时间：2026-06-11" in html
    assert "单号：CGXJ-260611-001" in html
    assert "佩文筛网<br>单价 / 总价" in html
    assert "捷阳五金<br>单价 / 总价" in html
    assert "lowest-cell" in html
    assert 'class="quote-unit-price">10.50</span>' in html
    assert 'class="quote-total-amount">21.00</span>' in html
    assert "10.50 / 21.00" not in html
    assert 'class="quote-cell selected-cell lowest-cell"' in html
    assert 'class="selected-material-row"' in html
    assert 'class="selected-supplier-cell"' in html
    assert ".selected-cell { background: #fff3bf;" in html
    assert "print-color-adjust: exact" in html
    assert "各供应商拟定合计" in html
    assert "佩文筛网: <strong>¥40.00</strong>" in html
    assert "捷阳五金: <strong>¥21.00</strong>" in html
    assert "拟定合计：¥61.00" in html
    assert "申请人签字" in html
    assert "材料员签字" in html
    assert "项目负责人签字" in html
    assert "项目管理处" in html
    assert "总经理签字" not in html


def test_material_approval_owner_can_approve_inquiry(client, test_db):
    cursor = test_db.cursor()
    cursor.execute("INSERT INTO roles (role_name, permissions) VALUES (?, ?)", ("材料审批负责人", ""))
    role_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO users (username, password, real_name, role_id, is_active, create_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        ("approver", "x", "审批负责人", role_id, 1, "2026-01-01 09:00:00"),
    )
    approver_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiries (
            inquiry_no, inquiry_date, applicant_id, total_amount,
            is_below_library_price, approval_status, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        ("XJ-APPROVE-001", "2026-06-06", approver_id, 0, 0, "待审批", "2026-06-06 10:00:00"),
    )
    inquiry_id = cursor.lastrowid
    test_db.commit()

    with client.session_transaction() as sess:
        sess["user"] = {
            "id": approver_id,
            "username": "approver",
            "real_name": "审批负责人",
            "role_name": "材料审批负责人",
        }

    response = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "reject", "remark": "测试审批"},
    )

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is True


def test_admin_and_material_approval_owner_can_approve_unpublished_quote_inquiry(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    admin_id = seed_role_user(cursor, "系统管理员", "admin_quote", "系统管理员")
    approver_id = seed_role_user(cursor, "材料审批负责人", "quote_approver", "审批负责人")

    admin_inquiry_id = seed_inquiry(cursor, "XJ-QUOTE-DRAFT-ADMIN", approver_id, status="报价未发布")
    approver_inquiry_id = seed_inquiry(cursor, "XJ-QUOTE-DRAFT-APPROVER", admin_id, status="报价未发布")
    test_db.commit()

    set_session_user(client, admin_id, "admin_quote", "系统管理员", "系统管理员")
    admin_response = client.post(
        f"/api/purchase-inquiries/{admin_inquiry_id}/approve",
        json={"action": "manager", "remark": "报价未发布管理员审批"},
    )

    assert admin_response.status_code == 200
    assert json.loads(admin_response.data)["success"] is True

    set_session_user(client, approver_id, "quote_approver", "审批负责人", "材料审批负责人")
    approver_response = client.post(
        f"/api/purchase-inquiries/{approver_inquiry_id}/approve",
        json={"action": "manager", "remark": "报价未发布负责人审批"},
    )

    assert approver_response.status_code == 200
    assert json.loads(approver_response.data)["success"] is True

    cursor.execute(
        "SELECT inquiry_no, approval_status FROM purchase_inquiries WHERE id IN (?, ?) ORDER BY inquiry_no",
        (admin_inquiry_id, approver_inquiry_id),
    )
    rows = cursor.fetchall()
    assert [row["approval_status"] for row in rows] == ["已同意", "已同意"]


def test_admin_cannot_approve_gx_project_inquiry(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    admin_id, _lei_id, _tan_id, _wang_id = seed_special_approval_users(cursor)
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-001", "GX项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-GX-ADMIN", admin_id, project_id=project_id)
    test_db.commit()
    set_session_user(client, admin_id, "admin", "管理员", "系统管理员")

    response = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "admin审批"},
    )

    data = json.loads(response.data)
    assert data["success"] is False
    assert "雷克峰和谭香" in data["message"]
    cursor.execute("SELECT approval_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == "待审批"


def test_gx_project_inquiry_requires_leikefeng_and_tanxiang_before_agreed(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    admin_id, lei_id, tan_id, _wang_id = seed_special_approval_users(cursor)
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-002", "GX项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-GX-DOUBLE", admin_id, project_id=project_id)
    test_db.commit()

    set_session_user(client, lei_id, "leikefeng", "雷克峰", "材料审批负责人")
    first = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "雷克峰审批"},
    )
    assert json.loads(first.data)["success"] is True
    cursor.execute("SELECT approval_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == "待审批"

    set_session_user(client, tan_id, "tanxiang", "谭香", "材料审批负责人")
    second = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "谭香审批"},
    )
    assert json.loads(second.data)["success"] is True
    cursor.execute("SELECT approval_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == "已同意"


def test_wanglihua_submitted_inquiry_requires_special_double_approval(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    _admin_id, lei_id, tan_id, wang_id = seed_special_approval_users(cursor)
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("KM-001", "普通项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-WLH-DOUBLE", wang_id, project_id=project_id)
    test_db.commit()

    set_session_user(client, tan_id, "tanxiang", "谭香", "材料审批负责人")
    first = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "谭香审批"},
    )
    assert json.loads(first.data)["success"] is True
    cursor.execute("SELECT approval_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == "待审批"

    set_session_user(client, lei_id, "leikefeng", "雷克峰", "材料审批负责人")
    second = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "雷克峰审批"},
    )
    assert json.loads(second.data)["success"] is True
    cursor.execute("SELECT approval_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    assert cursor.fetchone()[0] == "已同意"


def test_special_inquiry_rejects_duplicate_or_other_approver(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    admin_id, lei_id, _tan_id, _wang_id = seed_special_approval_users(cursor)
    other_id = seed_role_user(cursor, "材料审批负责人", "other_approver", "其他审批人")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-003", "GX项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-GX-DUP", admin_id, project_id=project_id)
    test_db.commit()

    set_session_user(client, lei_id, "leikefeng", "雷克峰", "材料审批负责人")
    first = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "第一次"},
    )
    assert json.loads(first.data)["success"] is True

    duplicate = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "重复审批"},
    )
    duplicate_data = json.loads(duplicate.data)
    assert duplicate_data["success"] is False
    assert "已审批过" in duplicate_data["message"]

    set_session_user(client, other_id, "other_approver", "其他审批人", "材料审批负责人")
    other = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "其他人审批"},
    )
    other_data = json.loads(other.data)
    assert other_data["success"] is False
    assert "雷克峰和谭香" in other_data["message"]


def test_gx_inquiry_creates_gx_material_for_cross_region_item(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    ensure_stock_in_project_id_column(cursor)
    _admin_id, lei_id, tan_id, _wang_id = seed_special_approval_users(cursor)
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("GX-004", "广西项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (
            material_code, material_name, specification, detail_spec, unit_id,
            tax_price, tax_exempt_price, freight, remark, inventory_min,
            inventory_max, create_time, tax_rate, project_id, weight,
            brand, is_national_standard, is_cash_price, cash_price, cash_tax_price
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "CDLX00001", "跨区材料", "DN20", "原详细规格", None,
            10, 9.9, 0, "", 0,
            0, "2026-01-01 09:00:00", 0.01, None, 0,
            "原品牌", 0, 0, 0, 0,
        ),
    )
    material_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-GX-MAT", lei_id, project_id=project_id)
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_items (
            inquiry_id, material_id, quantity, tax_rate, detail_spec, brand, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (inquiry_id, material_id, 2, 0.01, "广西详细规格", "广西品牌", "2026-06-06 10:00:00"),
    )
    item_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("广西报价供应商",))
    supplier_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_lowest, is_selected, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, supplier_id, 12, 11.88, 0.01, 24, 1, 1, "2026-06-06 10:00:00"),
    )
    test_db.commit()

    set_session_user(client, lei_id, "leikefeng", "雷克峰", "材料审批负责人")
    first = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "雷克峰审批"},
    )
    assert json.loads(first.data)["success"] is True

    set_session_user(client, tan_id, "tanxiang", "谭香", "材料审批负责人")
    second = client.post(
        f"/api/purchase-inquiries/{inquiry_id}/approve",
        json={"action": "manager", "remark": "谭香审批"},
    )
    assert json.loads(second.data)["success"] is True

    item_row = test_db.execute(
        "SELECT material_id FROM purchase_inquiry_items WHERE id = ?",
        (item_id,),
    ).fetchone()
    new_material = test_db.execute(
        "SELECT material_code, project_id, detail_spec, brand, tax_price FROM materials WHERE id = ?",
        (item_row["material_id"],),
    ).fetchone()
    assert item_row["material_id"] != material_id
    assert new_material["material_code"] == "GXLX00001"
    assert new_material["project_id"] == project_id
    assert new_material["detail_spec"] == "广西详细规格"
    assert new_material["brand"] == "广西品牌"
    assert new_material["tax_price"] == 12


def test_export_supplier_orders_shows_city_code_and_project_name(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    admin_id = seed_role_user(cursor, "系统管理员", "admin", "管理员")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("CD-QUOTE", "成都报价项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("测试供应商",))
    supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name) VALUES (?)", ("个",))
    unit_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, specification, unit_id)
        VALUES (?, ?, ?, ?)
        """,
        ("CDLX00001", "测试材料", "DN20", unit_id),
    )
    material_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-CD-EXPORT", admin_id, status="已同意", project_id=project_id)
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_items (inquiry_id, material_id, quantity, create_time)
        VALUES (?, ?, ?, ?)
        """,
        (inquiry_id, material_id, 3, "2026-06-06 10:00:00"),
    )
    item_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_lowest, is_selected, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, supplier_id, 12, 10.62, 0.13, 36, 1, 1, "2026-06-06 10:00:00"),
    )
    test_db.commit()
    set_session_user(client, admin_id, "admin", "管理员", "系统管理员")

    response = client.get(f"/api/purchase-inquiries/{inquiry_id}/export-supplier-orders")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook["测试供应商"]

    assert sheet["A4"].value == "项目：成都 / CD-QUOTE / 成都报价项目"


def test_export_draft_inquiry_xlsx_matches_quote_template(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    clerk_id = seed_role_user(cursor, "材料员", "clerk_export", "材料员")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("CD-DRAFT", "成都草稿项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("测试供应商",))
    supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name) VALUES (?)", ("个",))
    unit_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (material_code, material_name, specification, unit_id)
        VALUES (?, ?, ?, ?)
        """,
        ("CDLX00002", "测试材料", "DN25", unit_id),
    )
    material_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-DRAFT-EXPORT", clerk_id, status="草稿", project_id=project_id)
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_items (
            inquiry_id, material_id, quantity, tax_rate, is_national_standard,
            detail_spec, brand, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (inquiry_id, material_id, 5, 0.01, 1, "加厚", "测试品牌", "2026-06-06 10:00:00"),
    )
    item_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_lowest, is_selected, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, supplier_id, 0, 0, 0.01, 0, 0, 0, "2026-06-06 10:00:00"),
    )
    test_db.commit()
    set_session_user(client, clerk_id, "clerk_export", "材料员", "材料员")

    response = client.get(f"/api/purchase-inquiries/draft/{inquiry_id}/export-quote-sheet")

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "filename*=UTF-8''" in response.headers["Content-Disposition"]
    workbook = load_workbook(BytesIO(response.data), data_only=False)
    sheet = workbook.active

    assert sheet.title == "询价表"
    assert sheet["A1"].value == "零星材采购比价表"
    assert sheet["A2"].value == "项目名称：成都 / CD-DRAFT / 成都草稿项目"
    assert sheet["I2"].value == "时间：2026-06-06"
    assert [sheet.cell(3, col).value for col in range(1, 11)] == [
        "序号", "材料名称", "规格型号", "详细规格", "品牌", "是否国标", "单位", "数量", "测试供应商单价1%专票", "测试供应商总价"
    ]
    assert [sheet.cell(4, col).value for col in range(1, 9)] == [
        1, "测试材料", "DN25", "加厚", "测试品牌", "是", "个", 5
    ]
    assert sheet["J4"].value == '=IF(I4="","",I4*$H4)'


def test_import_draft_quote_sheet_parses_exported_template(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    clerk_id = seed_role_user(cursor, "材料员", "clerk_import", "材料员")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("CD-IMPORT", "成都导入项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("导入供应商",))
    supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name) VALUES (?)", ("个",))
    unit_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO materials (
            material_code, material_name, specification, detail_spec, brand,
            unit_id, tax_price, cash_price
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        ("CDLX00009", "导入材料", "DN40", "加厚款", "导入品牌", unit_id, 10, 9),
    )
    material_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-DRAFT-IMPORT", clerk_id, status="草稿", project_id=project_id)
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_items (
            inquiry_id, material_id, quantity, tax_rate, is_national_standard,
            detail_spec, brand, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (inquiry_id, material_id, 5, 0.01, 1, "加厚款", "导入品牌", "2026-06-06 10:00:00"),
    )
    item_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_lowest, is_selected, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, supplier_id, 0, 0, 0.13, 0, 0, 0, "2026-06-06 10:00:00"),
    )
    test_db.commit()
    set_session_user(client, clerk_id, "clerk_import", "材料员", "材料员")

    export_response = client.get(f"/api/purchase-inquiries/draft/{inquiry_id}/export-quote-sheet")
    workbook = load_workbook(BytesIO(export_response.data), data_only=False)
    sheet = workbook.active
    sheet["H4"] = 7
    sheet["I4"] = 12.5
    uploaded = BytesIO()
    workbook.save(uploaded)
    uploaded.seek(0)

    response = client.post(
        f"/api/purchase-inquiries/draft/{inquiry_id}/import-quote-sheet",
        data={"file": (uploaded, "询价表.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is True
    assert data["warnings"] == []
    assert len(data["items"]) == 1
    item = data["items"][0]
    assert item["material_id"] == material_id
    assert item["material_code"] == "CDLX00009"
    assert item["material_name"] == "导入材料"
    assert item["specification"] == "DN40"
    assert item["detail_spec"] == "加厚款"
    assert item["brand"] == "导入品牌"
    assert item["unit_name"] == "个"
    assert item["quantity"] == 7
    assert item["is_national_standard"] == 1
    assert item["unmatched_material"] is False
    assert item["quotes"][0]["supplier_id"] == supplier_id
    assert item["quotes"][0]["supplier_name"] == "导入供应商"
    assert item["quotes"][0]["tax_price"] == 12.5
    assert item["quotes"][0]["tax_rate"] == 0.13


def test_export_draft_inquiry_xlsx_tolerates_legacy_optional_columns(client, test_db):
    cursor = test_db.cursor()
    create_inquiry_delete_tables(cursor)
    ensure_project_id_column(cursor)
    for table, column in [
        ("purchase_inquiry_items", "brand"),
        ("purchase_inquiry_items", "detail_spec"),
        ("purchase_inquiry_items", "is_national_standard"),
        ("materials", "brand"),
        ("materials", "detail_spec"),
        ("materials", "is_national_standard"),
    ]:
        try:
            cursor.execute(f"ALTER TABLE {table} DROP COLUMN {column}")
        except Exception:
            pass

    clerk_id = seed_role_user(cursor, "材料员", "legacy_clerk", "材料员")
    cursor.execute(
        "INSERT INTO projects (project_code, project_name, create_time) VALUES (?, ?, ?)",
        ("CD-LEGACY", "成都旧库项目", "2026-01-01 09:00:00"),
    )
    project_id = cursor.lastrowid
    cursor.execute("INSERT INTO suppliers (supplier_name) VALUES (?)", ("旧库供应商",))
    supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name) VALUES (?)", ("米",))
    unit_id = cursor.lastrowid
    cursor.execute(
        "INSERT INTO materials (material_code, material_name, specification, unit_id) VALUES (?, ?, ?, ?)",
        ("CDLX00003", "旧库材料", "DN32", unit_id),
    )
    material_id = cursor.lastrowid
    inquiry_id = seed_inquiry(cursor, "XJ-DRAFT-LEGACY", clerk_id, status="草稿", project_id=project_id)
    cursor.execute(
        "INSERT INTO purchase_inquiry_items (inquiry_id, material_id, quantity, create_time) VALUES (?, ?, ?, ?)",
        (inquiry_id, material_id, 8, "2026-06-06 10:00:00"),
    )
    item_id = cursor.lastrowid
    cursor.execute(
        """
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_lowest, is_selected, create_time
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (item_id, supplier_id, 0, 0, 0.01, 0, 0, 0, "2026-06-06 10:00:00"),
    )
    test_db.commit()
    set_session_user(client, clerk_id, "legacy_clerk", "材料员", "材料员")

    response = client.get(f"/api/purchase-inquiries/draft/{inquiry_id}/export-quote-sheet")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data), data_only=False)
    sheet = workbook.active
    assert sheet["B4"].value == "旧库材料"
    assert sheet["C4"].value == "DN32"
    assert sheet["F4"].value == "否"


def test_export_draft_inquiry_does_not_require_openpyxl_before_auth(client, monkeypatch):
    original_import = builtins.__import__

    def reject_openpyxl(name, *args, **kwargs):
        if name.startswith("openpyxl"):
            raise ModuleNotFoundError("No module named 'openpyxl'")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", reject_openpyxl)

    response = client.get("/api/purchase-inquiries/draft/1/export-quote-sheet")

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data["success"] is False
    assert "登录" in data["message"]
