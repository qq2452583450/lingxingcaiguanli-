from io import BytesIO

from openpyxl import load_workbook


def test_export_supplier_orders_includes_detail_spec_between_spec_and_brand(client, test_db):
    cursor = test_db.cursor()
    now = '2026-06-11 10:00:00'

    cursor.execute("INSERT INTO suppliers (supplier_name, create_time) VALUES (?, ?)", ('测试供应商', now))
    supplier_id = cursor.lastrowid
    cursor.execute("INSERT INTO units (unit_name, unit_code) VALUES (?, ?)", ('件', 'PCS'))
    unit_id = cursor.lastrowid
    cursor.execute("""
        INSERT INTO materials (material_code, material_name, specification, detail_spec, brand, unit_id, create_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, ('CL-001', '钢管', 'DN50', '6米/根', '宝钢', unit_id, now))
    material_id = cursor.lastrowid
    cursor.execute("""
        INSERT INTO purchase_inquiries (inquiry_no, inquiry_date, total_amount, create_time)
        VALUES (?, ?, ?, ?)
    """, ('CGXJ-260611-001', '2026-06-11', 246, now))
    inquiry_id = cursor.lastrowid
    cursor.execute("""
        INSERT INTO purchase_inquiry_items (
            inquiry_id, material_id, quantity, library_price, detail_spec, brand, create_time
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (inquiry_id, material_id, 3, 10, '询价详细规格', '询价品牌', now))
    item_id = cursor.lastrowid
    cursor.execute("""
        INSERT INTO purchase_inquiry_quotes (
            item_id, supplier_id, tax_price, tax_exempt_price, tax_rate,
            total_amount, is_selected, create_time
        ) VALUES (?, ?, ?, ?, ?, ?, 1, ?)
    """, (item_id, supplier_id, 0.7345, 0.65, 0.13, 2.2035, now))
    cursor.execute("""
        INSERT INTO purchase_inquiry_supplier_freights (
            inquiry_id, supplier_id, tax_freight, tax_exempt_freight, tax_rate, remark, create_time, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (inquiry_id, supplier_id, 20, 17.7, 0.13, '整车配送', now, now))
    test_db.commit()

    response = client.get(f'/api/purchase-inquiries/{inquiry_id}/export-supplier-orders')

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    sheet = workbook['测试供应商']

    headers = [sheet.cell(row=6, column=col).value for col in range(1, 11)]
    assert headers == ['序号', '材料编码', '材料名称', '规格型号', '详细规格', '品牌', '单位', '数量', '含税单价', '金额']
    assert sheet.cell(row=7, column=4).value == 'DN50'
    assert sheet.cell(row=7, column=5).value == '询价详细规格'
    assert sheet.cell(row=7, column=6).value == '询价品牌'
    assert sheet.cell(row=7, column=9).value == 0.7345
    assert sheet.cell(row=7, column=10).value == 2.2035
    assert sheet.cell(row=7, column=8).number_format == 'General'
    assert sheet.cell(row=7, column=9).number_format == '#,##0.00'
    assert sheet.cell(row=7, column=10).number_format == '#,##0.00'
    assert sheet.cell(row=8, column=10).value == 2.2035
    assert sheet.cell(row=9, column=1).value == '运费'
    assert sheet.cell(row=9, column=10).value == 20
    assert sheet.cell(row=10, column=1).value == '到货总价'
    assert sheet.cell(row=10, column=10).value == 22.2035
