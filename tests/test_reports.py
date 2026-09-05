"""Financial reconciliation, project boundaries and full-export report regressions."""

from io import BytesIO
import sqlite3

import pytest
from flask import Flask
from openpyxl import load_workbook

from blueprints.reports import report_bp
from services.report_service import build_report, filters_from_args


@pytest.fixture
def report_db():
    c = sqlite3.connect(':memory:')
    c.row_factory = sqlite3.Row
    c.executescript('''
        CREATE TABLE roles(id INTEGER PRIMARY KEY,role_name TEXT);
        CREATE TABLE users(id INTEGER PRIMARY KEY,role_id INTEGER,is_active INTEGER);
        CREATE TABLE user_projects(user_id INTEGER,project_id INTEGER);
        CREATE TABLE projects(id INTEGER PRIMARY KEY,project_name TEXT);
        CREATE TABLE suppliers(id INTEGER PRIMARY KEY,supplier_name TEXT);
        CREATE TABLE units(id INTEGER PRIMARY KEY,unit_name TEXT);
        CREATE TABLE materials(id INTEGER PRIMARY KEY,material_name TEXT,material_code TEXT,
            specification TEXT,detail_spec TEXT,unit_id INTEGER);
        CREATE TABLE purchase_inquiries(id INTEGER PRIMARY KEY,inquiry_no TEXT,inquiry_date TEXT,
            project_id INTEGER,approval_status TEXT,selected_supplier_id INTEGER,total_amount REAL);
        CREATE TABLE purchase_inquiry_items(id INTEGER PRIMARY KEY,inquiry_id INTEGER,material_id INTEGER,
            quantity REAL,selected_quote_id INTEGER,detail_spec TEXT);
        CREATE TABLE purchase_inquiry_quotes(id INTEGER PRIMARY KEY,item_id INTEGER,supplier_id INTEGER,
            tax_price REAL,is_selected INTEGER,is_lowest INTEGER);
        CREATE TABLE purchase_inquiry_details(id INTEGER PRIMARY KEY,inquiry_id INTEGER,
            material_id INTEGER,supplier_id INTEGER,this_price REAL);
        CREATE TABLE purchase_inquiry_supplier_freights(inquiry_id INTEGER,supplier_id INTEGER,tax_freight REAL);
        CREATE TABLE stock_in_orders(id INTEGER PRIMARY KEY,order_no TEXT,in_time TEXT,
            project_id INTEGER,warehouse_id INTEGER,supplier_id INTEGER);
        CREATE TABLE stock_out_orders(id INTEGER PRIMARY KEY,order_no TEXT,out_time TEXT,
            project_id INTEGER,warehouse_id INTEGER);
        CREATE TABLE stock_in_details(id INTEGER PRIMARY KEY,order_id INTEGER,material_id INTEGER,
            quantity REAL,unit_price REAL,amount REAL,supplier_id INTEGER);
        CREATE TABLE stock_out_details(id INTEGER PRIMARY KEY,order_id INTEGER,material_id INTEGER,
            quantity REAL,unit_price REAL,amount REAL);
        CREATE TABLE warehouses(id INTEGER PRIMARY KEY,warehouse_name TEXT);
        CREATE TABLE inventory(id INTEGER PRIMARY KEY,material_id INTEGER,warehouse_id INTEGER,
            quantity REAL,unit_price REAL,update_time TEXT);
        CREATE TABLE base_inventory(id INTEGER PRIMARY KEY,material_id INTEGER,material_name TEXT,
            specification TEXT,detail_spec TEXT,unit_name TEXT,region TEXT,source_project_id INTEGER,
            quantity REAL,unit_price REAL,update_time TEXT);
        CREATE TABLE base_inventory_transfers(id INTEGER PRIMARY KEY,transfer_no TEXT,batch_no TEXT,
            transfer_time TEXT,project_id INTEGER,material_name TEXT,specification TEXT,detail_spec TEXT,
            unit_name TEXT,quantity REAL,depreciated_unit_price REAL,freight REAL);
        CREATE TABLE petty_cash_loans(id INTEGER PRIMARY KEY,project_id INTEGER,loan_no TEXT);
        CREATE TABLE petty_cash_usages(id INTEGER PRIMARY KEY,loan_id INTEGER,usage_no TEXT,use_date TEXT,
            expense_type TEXT,material_name TEXT,supplier_name TEXT,handler TEXT,amount REAL,
            invoice_amount REAL,is_reimbursed INTEGER,reimbursed_at TEXT);
    ''')
    c.executemany('INSERT INTO roles VALUES (?,?)', [(1,'系统管理员'),(2,'材料员'),(3,'供应商')])
    c.executemany('INSERT INTO users VALUES (?,?,1)', [(1,1),(2,2),(3,3),(4,2)])
    c.execute('INSERT INTO user_projects VALUES (2,1)')
    c.executemany('INSERT INTO projects VALUES (?,?)', [(1,'甲项目'),(2,'乙项目')])
    c.executemany('INSERT INTO suppliers VALUES (?,?)', [(11,'甲供应商'),(22,'乙供应商')])
    c.executemany('INSERT INTO units VALUES (?,?)', [(1,'根'),(2,'米')])
    c.executemany('INSERT INTO materials VALUES (?,?,?,?,?,?)', [(1,'钢管','M1','规格一','壁厚2mm',1),(2,'钢管','M2','规格二','壁厚3mm',2)])
    c.executemany('INSERT INTO purchase_inquiries VALUES (?,?,?,?,?,?,?)', [
        (1,'P1','2026-08-31 23:59:59',1,'已同意',None,370),
        (2,'P2','2026-09-01',1,'已同意',None,900),
        (3,'P3','2026-08-15',2,'已同意',None,210),
        (4,'P4','2026-08-15',1,'待审批',None,1000),
    ])
    c.executemany('INSERT INTO purchase_inquiry_items VALUES (?,?,?,?,?,?)', [
        (1,1,1,10,11,'壁厚2mm'),(2,1,2,5,22,'壁厚3mm'),(3,1,1,2,11,'壁厚2mm'),
        (4,2,1,9,11,''),(5,3,1,2,22,''),(6,4,1,10,11,''),
    ])
    # Quote ID 11 belongs to supplier 22; selected_quote_id=11 must select supplier 11.
    c.executemany('INSERT INTO purchase_inquiry_quotes VALUES (?,?,?,?,?,?)', [
        (11,1,22,1,0,1),(12,1,11,10,1,0),
        (13,2,22,40,1,1),(14,2,11,35,0,0),(15,3,11,10,1,1),
        (16,4,11,100,1,1),(17,5,22,100,1,1),(18,6,11,100,1,1),
    ])
    c.executemany('INSERT INTO purchase_inquiry_supplier_freights VALUES (?,?,?)', [(1,11,30),(1,22,20),(3,22,10)])
    c.commit()
    yield c
    c.close()


def report(c, kind='purchase', scope=None, **filters):
    f = filters_from_args({'kind':kind,'start_date':'2026-08-01','end_date':'2026-08-31',**filters})
    return build_report(c,f,scope)


def dataset(result, key):
    return next(t['rows'] for t in result['tables'] if t['key']==key)


def summary(result):
    return {label:value for label,value,_ in result['summary']}


def test_selected_supplier_counts_freight_dates_and_group_reconciliation(report_db):
    result = report(report_db)
    assert summary(result)['选定合计'] == 580
    assert summary(result)['询价单数'] == 2  # Sep 1 and pending excluded.
    assert summary(result)['选定运费'] == 60
    by_supplier = {r['supplier_id']: r for r in dataset(result,'supplier_amounts')}
    assert by_supplier[11]['count'] == 1  # two materials, one inquiry
    assert by_supplier[11]['amount'] == 150
    assert by_supplier[22]['count'] == 2
    assert by_supplier[22]['amount'] == 430
    assert sum(r['amount'] for r in dataset(result,'projects')) == 580
    assert sum(r['goods'] for r in dataset(result,'material_amounts')) + sum(r['freight'] for r in dataset(result,'freights')) == 580
    assert len(dataset(result,'material_quantities')) == 3  # separate details and units
    assert dataset(result,'checks') == []


def test_supplier_filter_keeps_freight_once_and_material_filter_omits_freight(report_db):
    assert summary(report(report_db,supplier_id='11'))['选定合计'] == 150
    assert summary(report(report_db,keyword='规格一'))['选定合计'] == 320
    assert summary(report(report_db,unit='根'))['选定运费'] == 0
    assert summary(report(report_db,keyword='不存在'))['选定合计'] == 0


def test_scope_excludes_unbound_projects_and_empty_scope_sees_no_rows(report_db):
    assert summary(report(report_db,scope=[1]))['选定合计'] == 370
    assert summary(report(report_db,scope=[]))['选定合计'] == 0
    with pytest.raises(PermissionError):
        report(report_db,scope=[1],project_id='2')


def test_header_supplier_and_ambiguous_selection_are_not_double_counted(report_db):
    report_db.execute('UPDATE purchase_inquiries SET selected_supplier_id=11 WHERE id=1')
    assert summary(report(report_db,scope=[1]))['选定合计'] == 325  # 120+175+30
    report_db.execute('UPDATE purchase_inquiry_items SET selected_quote_id=NULL WHERE id=1')
    report_db.execute('UPDATE purchase_inquiries SET selected_supplier_id=NULL WHERE id=1')
    report_db.execute('UPDATE purchase_inquiry_quotes SET is_selected=1 WHERE id=11')
    result=report(report_db,scope=[1])
    assert len(dataset(result,'details')) == 2
    assert any('多个' in r['note'] for r in dataset(result,'checks'))


def test_legacy_missing_quantity_and_empty_approved_orders_remain_visible(report_db):
    report_db.execute("INSERT INTO purchase_inquiries VALUES (5,'OLD','2026-08-01',1,'已同意',NULL,700)")
    report_db.execute('INSERT INTO purchase_inquiry_details VALUES (1,5,1,11,700)')
    report_db.execute("INSERT INTO purchase_inquiries VALUES (6,'EMPTY','2026-08-01',1,'已同意',NULL,10)")
    result=report(report_db,scope=[1])
    assert summary(result)['询价单数'] == 3
    assert sum(r['count'] for r in dataset(result,'projects')) == 3
    assert any('缺少数量' in r['note'] for r in dataset(result,'checks'))
    assert any('缺少材料明细' in r['note'] for r in dataset(result,'checks'))


def test_stock_scope_uses_same_warehouse_and_does_not_multiply_stock(report_db):
    report_db.executemany('INSERT INTO warehouses VALUES (?,?)', [(1,'一仓'),(2,'二仓')])
    report_db.executemany('INSERT INTO stock_in_orders VALUES (?,?,?,?,?,?)', [(1,'IN1','2026-08-02',1,1,11),(2,'IN2','2026-08-02',2,2,22),(3,'IN3','2026-08-03',1,1,11)])
    report_db.executemany('INSERT INTO stock_in_details VALUES (?,?,?,?,?,?,?)', [(1,1,1,3,10,30,11),(2,2,1,9,10,90,22),(3,3,1,2,10,20,11)])
    report_db.executemany('INSERT INTO inventory VALUES (?,?,?,?,?,?)', [(1,1,1,5,10,'2026-09-01'),(2,1,2,9,10,'2026-09-01')])
    assert summary(report(report_db,'stock_in',scope=[1]))['金额合计'] == 50
    result=report(report_db,'inventory',scope=[1])
    assert len(dataset(result,'details')) == 1
    assert summary(result)['金额合计'] == 50  # outside date range still current snapshot


def test_cash_includes_reimbursed_history_and_obeys_use_date(report_db):
    report_db.execute("INSERT INTO petty_cash_loans VALUES (1,1,'LOAN')")
    report_db.executemany('INSERT INTO petty_cash_usages VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', [
        (1,1,'U1','2026-08-31','运费',None,None,'甲',50,50,1,'2026-09-02'),
        (2,1,'U2','2026-08-01','加油费',None,None,'甲',100,0,0,None),
        (3,1,'U3','2026-09-01','加油费',None,None,'甲',999,0,0,None),
    ])
    s=summary(report(report_db,'cash',scope=[1]))
    assert (s['金额合计'],s['已报销支出'],s['未报销支出']) == (150,50,100)


def test_base_inventory_and_transfer_scope_use_source_and_destination_projects(report_db):
    report_db.executemany('INSERT INTO base_inventory VALUES (?,?,?,?,?,?,?,?,?,?,?)', [
        (1,None,'模板','甲','甲','张','云南',1,4,10,'2026-09-05'),
        (2,None,'模板','乙','乙','张','云南',2,8,10,'2026-09-05'),
    ])
    report_db.executemany('INSERT INTO base_inventory_transfers VALUES (?,?,?,?,?,?,?,?,?,?,?,?)', [
        (1,'T1','BATCH','2026-08-31 23:59:59',1,'模板','甲','甲','张',2,10,5),
        (2,'T2','BATCH','2026-08-31 23:59:59',1,'钢管','甲','甲','根',3,10,7),
        (3,'T3','BATCH2','2026-09-01',1,'钢管','甲','甲','根',9,10,7),
    ])
    result = report(report_db,'base_inventory',scope=[1])
    assert summary(result)['金额合计'] == 40
    result = report(report_db,'transfers',scope=[1])
    assert summary(result)['金额合计'] == 62
    assert summary(result)['业务笔数'] == 1
    assert dataset(result,'projects')[0]['freight'] == 12


def test_stock_out_keeps_record_amount_and_project_scope(report_db):
    report_db.executemany('INSERT INTO stock_out_orders VALUES (?,?,?,?,?)', [(1,'O1','2026-08-31 22:00',1,1),(2,'O2','2026-08-31',2,1)])
    report_db.executemany('INSERT INTO stock_out_details VALUES (?,?,?,?,?,?)', [(1,1,1,2,10,20),(2,2,1,3,10,30)])
    result=report(report_db,'stock_out',scope=[1])
    assert summary(result)['金额合计'] == 20
    assert dataset(result,'details')[0]['order_no'] == 'O1'


@pytest.fixture
def reports_client(report_db,monkeypatch):
    import blueprints.reports as module
    app=Flask(__name__)
    app.config.update(TESTING=True,SECRET_KEY='reports-test-only')
    app.register_blueprint(report_bp)
    monkeypatch.setattr(module,'get_db',lambda:report_db)
    return app.test_client()


def login(client,user_id):
    with client.session_transaction() as state:
        state['user']={'id':user_id,'role_name':'系统管理员'}  # spoofed role must not grant access


def test_api_auth_invalid_filters_and_project_export_boundaries(reports_client):
    assert reports_client.get('/api/reports').status_code == 401
    login(reports_client,3)
    assert reports_client.get('/api/reports').status_code == 403
    login(reports_client,2)
    assert reports_client.get('/api/reports/export?project_id=2').status_code == 403
    assert reports_client.get('/api/reports?start_date=2026-09-01&end_date=2026-08-01').status_code == 400
    assert reports_client.get('/api/reports?kind=invalid').status_code == 400
    options=reports_client.get('/api/reports/options').json['data']
    assert [r['id'] for r in options['projects']] == [1]


def test_api_pages_at_50_but_export_is_complete_and_formula_safe(report_db,reports_client):
    for i in range(100,160):
        report_db.execute('INSERT INTO purchase_inquiry_items VALUES (?,?,?,?,?,?)', (i,1,1,1,11,'export'))
        report_db.execute('INSERT INTO purchase_inquiry_quotes VALUES (?,?,?,?,?,?)', (i,i,11,10,1,1))
    report_db.execute("UPDATE suppliers SET supplier_name='=1+1' WHERE id=11")
    report_db.commit()
    login(reports_client,1)
    url='?start_date=2026-08-01&end_date=2026-08-31'
    response=reports_client.get('/api/reports'+url)
    assert response.status_code == 200
    data=response.json['data']
    assert len(dataset(data,'details')) == 50
    page2=reports_client.get('/api/reports'+url+'&page_details=2').json['data']
    assert len(dataset(page2,'details')) == 14
    export=reports_client.get('/api/reports/export'+url+'&page_details=2')
    assert export.status_code == 200
    book=load_workbook(BytesIO(export.data))
    sheet=book['选定采购明细']
    assert sheet.max_row == 65
    assert all(c.data_type != 'f' for row in sheet for c in row)
    assert any(c.value=='=1+1' for row in sheet for c in row)
    assert sum(row[-1].value for row in list(sheet)[1:]) == 1120
    assert sum(row[-1].value for row in list(book['选定供应商运费明细'])[1:]) == 60


@pytest.mark.parametrize('kind',['purchase','stock_in','stock_out','inventory','base_inventory','transfers','cash'])
def test_every_report_and_export_supports_empty_results(kind,reports_client):
    login(reports_client,4)  # no project grants
    url='?kind='+kind+'&start_date=2026-08-01&end_date=2026-08-31'
    response=reports_client.get('/api/reports'+url)
    assert response.status_code == 200
    assert all(not t['rows'] for t in response.json['data']['tables'])
    assert reports_client.get('/api/reports/export'+url).status_code == 200
