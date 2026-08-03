"""Exam center API blueprint."""

import sqlite3
import time
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file, session

from helpers import get_db
from services.exam_service import (
    can_manage_exam,
    can_take_exam,
    clear_practice_draft,
    DAILY_PRACTICE_QUESTION_COUNT,
    delete_exam_attempt,
    get_attempt,
    get_attempt_review,
    list_attendance_calendar,
    get_current_exam_paper,
    get_daily_practice_status,
    get_exam_paper,
    get_formal_exam_pool_status,
    get_paper_questions,
    get_practice_draft,
    get_random_formal_exam_paper,
    get_wrong_practice_questions_for_retry,
    list_daily_checkins,
    get_random_practice_questions,
    list_practice_history,
    list_papers,
    list_pending_reviews,
    list_results,
    list_monthly_checkin_reports,
    list_retake_eligibilities,
    list_material_clerk_wrong_questions,
    list_wrong_practice_questions,
    record_practice_answers,
    retry_wrong_practice_answers,
    review_answer,
    save_practice_draft,
    start_attempt,
    start_retake_attempt,
    submit_retroactive_checkin,
    submit_attempt,
    FORMAL_EXAM_POOL_REQUIRED_COUNT,
    FORMAL_EXAM_POOL_SETTING_KEY,
)


exam_bp = Blueprint("exam", __name__, url_prefix="/api/exam")


def _current_user():
    return session.get("user") or {}


def _json_error(message, status_code=400):
    return jsonify({"success": False, "message": message}), status_code


def _ensure_exam_settings_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS exam_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )


def _clear_current_exam_setting(conn):
    last_error = None
    for _ in range(3):
        try:
            _ensure_exam_settings_table(conn)
            conn.execute(
                "DELETE FROM exam_settings WHERE key = ?",
                ("current_exam_paper_id",),
            )
            conn.commit()
            return
        except sqlite3.OperationalError as exc:
            last_error = exc
            conn.rollback()
            if "locked" in str(exc).lower() or "no such table" in str(exc).lower():
                time.sleep(0.2)
                continue
            raise
    raise last_error


def _require_exam_user():
    user = _current_user()
    if user and (can_take_exam(user) or can_manage_exam(user)):
        return user, None
    return None, _json_error("Permission denied", 403)


def _require_exam_manager():
    user = _current_user()
    if user and can_manage_exam(user):
        return user, None
    return None, _json_error("Permission denied", 403)


def _paper_exists(paper_id):
    return any(paper["id"] == paper_id for paper in list_papers())


def _sanitize_question(question):
    allowed_keys = {
        "id",
        "paper_id",
        "paper_title",
        "question_type",
        "order_no",
        "stem",
        "score",
        "options",
    }
    return {key: value for key, value in question.items() if key in allowed_keys}


def _questions_for_exam_flow(questions):
    return [_sanitize_question(question) for question in questions]


def _filters_from_request():
    filters = {}
    paper_id = request.args.get("paper_id", type=int)
    status = request.args.get("status", "")
    keyword = request.args.get("keyword", "")
    if paper_id:
        filters["paper_id"] = paper_id
    if status:
        filters["status"] = status
    if keyword:
        filters["keyword"] = keyword
    return filters


def _monthly_checkins_workbook(rows, month):
    from calendar import monthrange
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月度打卡对账"
    sheet.sheet_view.showGridLines = False

    try:
        year, month_number = map(int, (month or datetime.now().strftime("%Y-%m")).split("-"))
    except ValueError:
        year, month_number = datetime.now().year, datetime.now().month
    days_in_month = monthrange(year, month_number)[1]
    day_start_column = 4
    summary_headers = ["合格打卡", "补卡", "缺勤", "满勤标准", "满勤状态"]
    summary_start_column = day_start_column + days_in_month
    last_column = summary_start_column + len(summary_headers) - 1
    last_column_letter = get_column_letter(last_column)
    data_start_row = 6
    data_end_row = data_start_row + max(len(rows), 1) - 1

    navy_fill = PatternFill("solid", fgColor="1F4E78")
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    amber_fill = PatternFill("solid", fgColor="FFE699")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    gray_fill = PatternFill("solid", fgColor="E7E6E6")
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    centered = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells(f"A1:{last_column_letter}1")
    sheet["A1"] = "考试中心月度打卡对账单"
    sheet["A1"].font = Font(name="等线", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = navy_fill
    sheet["A1"].alignment = centered
    sheet.row_dimensions[1].height = 26

    summary_values = [
        ("A2", "对账月份"), ("B2", f"{year:04d}-{month_number:02d}"),
        ("D2", "材料员人数"), ("E2", f"=COUNTA(A{data_start_row}:A{data_end_row})"),
        ("G2", "满勤人数"), ("H2", f'=COUNTIF({get_column_letter(summary_start_column + 4)}{data_start_row}:{get_column_letter(summary_start_column + 4)}{data_end_row},"满勤")'),
        ("J2", "补卡总次数"), ("K2", f"=SUM({get_column_letter(summary_start_column + 1)}{data_start_row}:{get_column_letter(summary_start_column + 1)}{data_end_row})"),
        ("M2", "生成时间"), ("N2", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for cell_ref, value in summary_values:
        cell = sheet[cell_ref]
        cell.value = value
        cell.alignment = centered
        cell.border = thin_border
        if cell_ref[0] in {"A", "D", "G", "J", "M"}:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
        else:
            cell.fill = blue_fill

    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    sheet.cell(row=4, column=1, value="人员信息")
    sheet.merge_cells(start_row=4, start_column=day_start_column, end_row=4, end_column=summary_start_column - 1)
    sheet.cell(row=4, column=day_start_column, value="日期")
    sheet.merge_cells(start_row=4, start_column=summary_start_column, end_row=4, end_column=last_column)
    sheet.cell(row=4, column=summary_start_column, value="月度统计")
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = centered
        cell.border = thin_border

    headers = ["姓名", "账号", "岗位"] + list(range(1, days_in_month + 1)) + summary_headers
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = centered
        cell.border = thin_border
    sheet.row_dimensions[5].height = 22

    status_styles = {
        "打卡": green_fill,
        "补卡": amber_fill,
        "未打卡": red_fill,
        "未达标": red_fill,
        "未来": gray_fill,
    }
    for row_index, report in enumerate(rows, start=data_start_row):
        statuses = []
        for day in report.get("days", []):
            if day.get("status") == "passed":
                statuses.append("补卡" if day.get("retroactive") else "打卡")
            elif day.get("status") == "failed":
                statuses.append("未达标")
            elif day.get("status") == "future":
                statuses.append("未来")
            else:
                statuses.append("未打卡")
        values = [
            report.get("real_name") or report.get("username") or "",
            report.get("username") or "",
            report.get("role_name") or "",
            *statuses,
            report.get("actual_days") or 0,
            report.get("retroactive_used") or 0,
            report.get("missing_days") or 0,
            report.get("expected_days") or 0,
            "满勤" if report.get("full_attendance") else "未满勤",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = centered
            cell.border = thin_border
            if day_start_column <= column < summary_start_column:
                cell.fill = status_styles.get(value, gray_fill)
            elif column == summary_start_column + 4:
                cell.fill = green_fill if value == "满勤" else red_fill
    sheet.auto_filter.ref = f"A5:{last_column_letter}{data_end_row}"
    sheet.freeze_panes = "D6"

    for column in range(1, last_column + 1):
        letter = get_column_letter(column)
        if column == 1:
            sheet.column_dimensions[letter].width = 12
        elif column in {2, 3, last_column}:
            sheet.column_dimensions[letter].width = 16
        elif day_start_column <= column < summary_start_column:
            sheet.column_dimensions[letter].width = 9
        else:
            sheet.column_dimensions[letter].width = 12

    legend_row = data_end_row + 2
    sheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=3)
    sheet.cell(row=legend_row, column=1, value="状态说明")
    sheet.cell(row=legend_row, column=1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=legend_row, column=1).fill = navy_fill
    sheet.cell(row=legend_row, column=1).alignment = centered
    for offset, label in enumerate(["打卡", "补卡", "未打卡", "未达标", "未来"], start=day_start_column):
        cell = sheet.cell(row=legend_row, column=offset, value=label)
        cell.fill = status_styles[label]
        cell.alignment = centered
        cell.border = thin_border

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:5"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _option_answer_text(question, answer_text):
    answer_text = str(answer_text or "").strip()
    if not answer_text:
        return "-"
    options = {str(option.get("key")): option.get("text") for option in question.get("options", [])}
    return "、".join(
        f"{key}. {options[key]}" if key in options else key
        for key in answer_text.split(",")
        if key
    ) or answer_text


def _material_clerk_wrong_questions_workbook(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "材料员错题集合"
    headers = [
        "序号", "来源", "试卷", "题型", "题目", "正确答案", "解析",
        "错题频次", "涉及材料员数", "材料员错答详情", "最近答错时间",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(rows, start=1):
        answer_details = "\n".join(
            f"{detail['user_name']}（{detail['source_label']}，{detail['wrong_count']}次）："
            f"{_option_answer_text(row, detail['answer_text'])}"
            for detail in row.get("answer_details", [])
        )
        sheet.append([
            index,
            row.get("source_labels", ""),
            row.get("paper_title", ""),
            row.get("question_type", ""),
            row.get("stem", ""),
            _option_answer_text(row, row.get("correct_answer")),
            row.get("reference_answer", ""),
            row.get("wrong_count", 0),
            row.get("clerk_count", 0),
            answer_details,
            row.get("created_at", ""),
        ])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in enumerate((8, 16, 24, 14, 42, 28, 42, 12, 14, 46, 20), start=1):
        sheet.column_dimensions[chr(64 + column)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@exam_bp.route("/summary", methods=["GET"])
def summary():
    user, denied = _require_exam_user()
    if denied:
        return denied

    papers = list_papers()
    return jsonify(
        {
            "success": True,
            "data": {
                "can_take": can_take_exam(user),
                "can_manage": can_manage_exam(user),
                "current_paper": get_current_exam_paper(),
                "formal_exam_pool": get_formal_exam_pool_status(),
                "paper_count": len(papers),
            },
        }
    )


@exam_bp.route("/practice/random", methods=["GET"])
def random_practice():
    user, denied = _require_exam_user()
    if denied:
        return denied

    limit = request.args.get("limit", default=DAILY_PRACTICE_QUESTION_COUNT, type=int)
    limit = max(1, min(limit or DAILY_PRACTICE_QUESTION_COUNT, 100))
    paper_id = request.args.get("paper_id", type=int)
    questions = get_random_practice_questions(limit=limit, paper_id=paper_id)
    return jsonify({"success": True, "data": _questions_for_exam_flow(questions)})


@exam_bp.route("/practice/submit", methods=["POST"])
def submit_practice():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    data = request.get_json(silent=True) or {}
    try:
        result = record_practice_answers(user["id"], data.get("answers") or {})
    except ValueError as exc:
        return _json_error(str(exc), 400)
    return jsonify({"success": True, "data": result})


@exam_bp.route("/practice/draft", methods=["GET"])
def practice_draft_detail():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    draft = get_practice_draft(user["id"])
    if draft:
        draft = dict(draft)
        draft["questions"] = _questions_for_exam_flow(draft.get("questions") or [])
    return jsonify({"success": True, "data": draft})


@exam_bp.route("/practice/draft", methods=["PUT"])
def save_practice_draft_route():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    data = request.get_json(silent=True) or {}
    try:
        draft = save_practice_draft(
            user["id"],
            data.get("question_ids") or [],
            data.get("answers") or {},
        )
    except ValueError as exc:
        return _json_error(str(exc), 400)
    draft["questions"] = _questions_for_exam_flow(draft.get("questions") or [])
    return jsonify({"success": True, "data": draft})


@exam_bp.route("/practice/draft", methods=["DELETE"])
def clear_practice_draft_route():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    clear_practice_draft(user["id"])
    return jsonify({"success": True})


@exam_bp.route("/practice/daily-status", methods=["GET"])
def practice_daily_status():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    return jsonify({"success": True, "data": get_daily_practice_status(user["id"])})


@exam_bp.route("/attendance/calendar", methods=["GET"])
def attendance_calendar():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    try:
        data = list_attendance_calendar(user["id"], request.args.get("month") or None)
    except ValueError as exc:
        return _json_error(str(exc), 400)
    return jsonify({"success": True, "data": data})


@exam_bp.route("/attendance/retroactive/submit", methods=["POST"])
def submit_retroactive_attendance():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    data = request.get_json(silent=True) or {}
    try:
        result = submit_retroactive_checkin(
            user["id"],
            data.get("target_date") or "",
            data.get("answers") or {},
        )
    except ValueError as exc:
        return _json_error(str(exc), 400)
    return jsonify({"success": True, "data": result})


@exam_bp.route("/practice/history", methods=["GET"])
def practice_history():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    limit = request.args.get("limit", default=100, type=int)
    limit = max(1, min(limit or 100, 500))
    return jsonify({"success": True, "data": list_practice_history(user["id"], limit=limit)})


@exam_bp.route("/practice/wrong", methods=["GET"])
def wrong_practice():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    limit = request.args.get("limit", default=100, type=int)
    limit = max(1, min(limit or 100, 500))
    return jsonify({"success": True, "data": list_wrong_practice_questions(user["id"], limit=limit)})


@exam_bp.route("/admin/practice/wrong-questions", methods=["GET"])
def admin_material_clerk_wrong_questions():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    limit = request.args.get("limit", default=100, type=int)
    limit = max(1, min(limit or 100, 500))
    return jsonify({"success": True, "data": list_material_clerk_wrong_questions(limit=limit)})


@exam_bp.route("/admin/practice/wrong-questions/export", methods=["GET"])
def export_material_clerk_wrong_questions():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    output = _material_clerk_wrong_questions_workbook(
        list_material_clerk_wrong_questions(limit=500)
    )
    return send_file(
        output,
        as_attachment=True,
        download_name="material-clerk-wrong-questions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@exam_bp.route("/practice/wrong/questions", methods=["GET"])
def wrong_practice_questions():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    limit = request.args.get("limit", default=100, type=int)
    limit = max(1, min(limit or 100, 500))
    questions = get_wrong_practice_questions_for_retry(user["id"], limit=limit)
    return jsonify({"success": True, "data": _questions_for_exam_flow(questions)})


@exam_bp.route("/practice/wrong/submit", methods=["POST"])
def submit_wrong_practice():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    data = request.get_json(silent=True) or {}
    try:
        result = retry_wrong_practice_answers(user["id"], data.get("answers") or {})
    except ValueError as exc:
        return _json_error(str(exc), 400)
    return jsonify({"success": True, "data": result})


@exam_bp.route("/papers", methods=["GET"])
def papers():
    _, denied = _require_exam_user()
    if denied:
        return denied

    return jsonify({"success": True, "data": list_papers()})


@exam_bp.route("/attempts", methods=["POST"])
def create_attempt():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    pool_status = get_formal_exam_pool_status()
    if pool_status["enabled"]:
        selected_paper = get_random_formal_exam_paper()
        if not selected_paper:
            return _json_error("Formal exam pool is not ready")
        paper_id = selected_paper["id"]
    else:
        current_paper = get_current_exam_paper()
        if not current_paper:
            return _json_error("No current exam paper is set")

        data = request.get_json(silent=True) or {}
        paper_id = data.get("paper_id")
        if paper_id is None:
            paper_id = current_paper["id"]
        try:
            paper_id = int(paper_id)
        except (TypeError, ValueError):
            return _json_error("Invalid paper_id")
        if paper_id != current_paper["id"]:
            return _json_error("Only the current exam paper can be attempted")
    if not _paper_exists(paper_id):
        return _json_error("Paper not found", 404)

    attempt_id = start_attempt(user["id"], paper_id)
    attempt = get_attempt(attempt_id)
    attempt["attempt_id"] = attempt_id
    return jsonify({"success": True, "attempt_id": attempt_id, "data": attempt})


@exam_bp.route("/attempts/<int:attempt_id>", methods=["GET"])
def attempt_detail(attempt_id):
    user, denied = _require_exam_user()
    if denied:
        return denied

    attempt = get_attempt(attempt_id)
    if not attempt:
        return _json_error("Attempt not found", 404)
    if attempt["user_id"] != user.get("id") and not can_manage_exam(user):
        return _json_error("Permission denied", 403)

    questions = get_paper_questions(attempt["paper_id"])
    attempt["questions"] = _questions_for_exam_flow(questions)
    return jsonify({"success": True, "data": attempt})


@exam_bp.route("/attempts/<int:attempt_id>/review", methods=["GET"])
def attempt_review(attempt_id):
    user, denied = _require_exam_user()
    if denied:
        return denied

    attempt = get_attempt(attempt_id)
    if not attempt:
        return _json_error("Attempt not found", 404)
    if attempt["user_id"] != user.get("id") and not can_manage_exam(user):
        return _json_error("Permission denied", 403)
    if attempt["status"] == "in_progress":
        return _json_error("Attempt has not been submitted", 400)

    return jsonify({"success": True, "data": get_attempt_review(attempt_id)})


@exam_bp.route("/attempts/<int:attempt_id>/submit", methods=["POST"])
def submit_attempt_route(attempt_id):
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    attempt = get_attempt(attempt_id)
    if not attempt:
        return _json_error("Attempt not found", 404)
    if attempt["user_id"] != user.get("id"):
        return _json_error("Permission denied", 403)

    data = request.get_json(silent=True) or {}
    try:
        submit_attempt(attempt_id, data.get("answers") or {})
    except ValueError as exc:
        return _json_error(str(exc), 400)

    return jsonify({"success": True, "data": get_attempt_review(attempt_id)})


@exam_bp.route("/results", methods=["GET"])
def results():
    user, denied = _require_exam_user()
    if denied:
        return denied

    filters = _filters_from_request()
    filters["viewer"] = user
    return jsonify({"success": True, "data": list_results(filters)})


@exam_bp.route("/retake/eligibilities", methods=["GET"])
def retake_eligibilities():
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    return jsonify(
        {
            "success": True,
            "data": list_retake_eligibilities(user_id=user["id"]),
        }
    )


@exam_bp.route("/retake/eligibilities/<int:eligibility_id>/start", methods=["POST"])
def start_retake(eligibility_id):
    user, denied = _require_exam_user()
    if denied:
        return denied
    if not can_take_exam(user):
        return _json_error("Permission denied", 403)

    try:
        attempt_id = start_retake_attempt(user["id"], eligibility_id)
    except ValueError as exc:
        status = 403 if str(exc) == "Permission denied" else 400
        return _json_error(str(exc), status)
    attempt = get_attempt(attempt_id)
    attempt["attempt_id"] = attempt_id
    return jsonify({"success": True, "attempt_id": attempt_id, "data": attempt})


@exam_bp.route("/admin/papers", methods=["GET"])
def admin_papers():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    return jsonify({"success": True, "data": list_papers()})


@exam_bp.route("/admin/current-paper", methods=["POST"])
def set_current_paper():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    data = request.get_json(silent=True) or {}
    try:
        paper_id = int(data.get("paper_id"))
    except (TypeError, ValueError):
        return _json_error("Invalid paper_id")
    paper = get_exam_paper(paper_id)
    if not paper:
        return _json_error("Paper not found", 404)
    if paper["source_type"] != "exam":
        return _json_error("只能设置正式考试卷为当前试卷")

    conn = get_db()
    _ensure_exam_settings_table(conn)
    conn.execute(
        "INSERT OR REPLACE INTO exam_settings (key, value) VALUES (?, ?)",
        ("current_exam_paper_id", str(paper_id)),
    )
    conn.execute(
        "DELETE FROM exam_settings WHERE key = ?",
        (FORMAL_EXAM_POOL_SETTING_KEY,),
    )
    conn.commit()
    return jsonify({"success": True, "data": get_current_exam_paper()})


@exam_bp.route("/admin/current-paper", methods=["DELETE"])
def clear_current_paper():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    conn = get_db()
    try:
        _clear_current_exam_setting(conn)
    except sqlite3.OperationalError as exc:
        return _json_error(f"取消当前试卷失败：{exc}", 500)
    return jsonify({"success": True, "data": None})


@exam_bp.route("/admin/formal-exam-pool", methods=["POST"])
def enable_formal_exam_pool():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    status = get_formal_exam_pool_status()
    if status["paper_count"] != FORMAL_EXAM_POOL_REQUIRED_COUNT:
        return _json_error("正式考试题池必须包含三套完整试卷")

    conn = get_db()
    _ensure_exam_settings_table(conn)
    conn.execute(
        "INSERT OR REPLACE INTO exam_settings (key, value) VALUES (?, ?)",
        (FORMAL_EXAM_POOL_SETTING_KEY, "1"),
    )
    conn.execute(
        "DELETE FROM exam_settings WHERE key = ?",
        ("current_exam_paper_id",),
    )
    conn.commit()
    return jsonify({"success": True, "data": get_formal_exam_pool_status()})


@exam_bp.route("/admin/formal-exam-pool", methods=["DELETE"])
def disable_formal_exam_pool():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    conn = get_db()
    _ensure_exam_settings_table(conn)
    conn.execute(
        "DELETE FROM exam_settings WHERE key = ?",
        (FORMAL_EXAM_POOL_SETTING_KEY,),
    )
    conn.commit()
    return jsonify({"success": True, "data": get_formal_exam_pool_status()})


@exam_bp.route("/admin/results", methods=["GET"])
def admin_results():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    return jsonify({"success": True, "data": list_results(_filters_from_request())})


@exam_bp.route("/admin/checkins", methods=["GET"])
def admin_checkins():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    target_date = request.args.get("date") or None
    return jsonify({"success": True, "data": list_daily_checkins(target_date)})


@exam_bp.route("/admin/checkins/monthly", methods=["GET"])
def admin_monthly_checkins():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    month = request.args.get("month") or None
    try:
        data = list_monthly_checkin_reports(month)
    except ValueError as exc:
        return _json_error(str(exc), 400)
    if request.args.get("format") == "json":
        return jsonify({"success": True, "data": data})

    export_month = month or (data[0]["month"] if data else "")
    output = _monthly_checkins_workbook(data, export_month)
    filename = f"monthly-checkins-{export_month or 'current'}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@exam_bp.route("/admin/retake/eligibilities", methods=["GET"])
def admin_retake_eligibilities():
    user, denied = _require_exam_manager()
    if denied:
        return denied

    return jsonify({"success": True, "data": list_retake_eligibilities(viewer=user)})


@exam_bp.route("/admin/attempts/<int:attempt_id>", methods=["DELETE"])
def admin_delete_attempt(attempt_id):
    _, denied = _require_exam_manager()
    if denied:
        return denied

    try:
        delete_exam_attempt(attempt_id)
    except ValueError as exc:
        return _json_error(str(exc), 404)
    return jsonify({"success": True})


@exam_bp.route("/admin/reviews", methods=["GET"])
def admin_reviews():
    _, denied = _require_exam_manager()
    if denied:
        return denied

    return jsonify({"success": True, "data": list_pending_reviews()})


@exam_bp.route("/admin/reviews/<int:answer_id>", methods=["POST"])
def admin_review_answer(answer_id):
    user, denied = _require_exam_manager()
    if denied:
        return denied

    data = request.get_json(silent=True) or {}
    try:
        final_score = float(data.get("final_score"))
    except (TypeError, ValueError):
        return _json_error("Invalid final_score")

    review_answer(answer_id, user["id"], final_score, data.get("comment") or "")
    return jsonify({"success": True})
