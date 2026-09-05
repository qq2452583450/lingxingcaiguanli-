"""Report queries and full-result Excel exports; no business data writes."""

from io import BytesIO
from math import ceil

from flask import Blueprint, jsonify, request, send_file, session

from helpers import get_db
from services.report_service import build_report, filters_from_args, project_scope, REPORT_NAMES


report_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


@report_bp.before_request
def authorize():
    if not session.get('user'):
        return jsonify(success=False, message='请先登录'), 401


@report_bp.errorhandler(ValueError)
def invalid_request(error):
    return jsonify(success=False, message=str(error)), 400


@report_bp.errorhandler(PermissionError)
def forbidden(error):
    return jsonify(success=False, message=str(error)), 403


@report_bp.route('/options')
def options():
    conn = get_db()
    scope = project_scope(conn, session['user'])
    projects = [dict(r) for r in conn.execute('SELECT id, project_name FROM projects ORDER BY project_name')
                if scope is None or r['id'] in scope]
    suppliers = [dict(r) for r in conn.execute('SELECT id, supplier_name FROM suppliers ORDER BY supplier_name')]
    units = [r[0] for r in conn.execute('SELECT unit_name FROM units ORDER BY unit_name')]
    return jsonify(success=True, data={'projects': projects, 'suppliers': suppliers, 'units': units,
                                      'reports': [{'key': k, 'title': v} for k, v in REPORT_NAMES.items()]})


def report_data():
    conn = get_db()
    scope = project_scope(conn, session['user'])
    f = filters_from_args(request.args)
    # All related reads, including export, see the same database snapshot.
    conn.execute('BEGIN')
    try:
        return build_report(conn, f, scope)
    finally:
        conn.rollback()


@report_bp.route('')
def query():
    data = report_data()
    for dataset in data['tables']:
        try:
            page = max(1, int(request.args.get('page_' + dataset['key'], '1')))
        except ValueError:
            raise ValueError('页码参数无效')
        total = len(dataset['rows'])
        pages = max(1, ceil(total / 50))
        page = min(page, pages)
        dataset.update(total=total, page=page, pages=pages, page_size=50)
        dataset['rows'] = dataset['rows'][(page - 1) * 50:page * 50]
    return jsonify(success=True, data=data)


def make_workbook(data):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    info = book.active
    info.title = '统计说明'
    info.append([data['title'], data['generated_at']])
    for key, value in data['filters'].items():
        info.append([{'kind': '报表类型', 'project_id': '项目编号', 'supplier_id': '供应商编号', 'keyword': '材料关键词', 'unit': '单位', 'start_date': '开始日期', 'end_date': '结束日期'}[key], str(value)])
    for label, value, _ in data['summary']:
        info.append([label, value])
    for note in data['notes']:
        info.append(['统计口径', note])
    for dataset in data['tables']:
        sheet = book.create_sheet(dataset['title'][:31])
        sheet.append([c['label'] for c in dataset['columns']])
        for row in dataset['rows']:
            sheet.append([row.get(c['key']) for c in dataset['columns']])
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(dataset['columns'], 1):
            sheet.column_dimensions[get_column_letter(index)].width = 22 if column['type'] == 'text' else 18
            if column['type'] != 'text':
                for cells in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                    cells[0].number_format = '#,##0.00' if column['type'] == 'money' else '#,##0.####'
    info.column_dimensions['A'].width = 24
    info.column_dimensions['B'].width = 95
    for sheet in book:
        sheet.freeze_panes = 'A2'
        for row in sheet:
            for cell in row:
                # Names and keywords from business data are literal text, not Excel formulas.
                if isinstance(cell.value, str):
                    cell.data_type = 's'
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if cell.row == 1:
                    cell.fill = PatternFill('solid', fgColor='244B78')
                    cell.font = Font(name='微软雅黑', color='FFFFFF', bold=True)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
    output = BytesIO()
    book.save(output)
    output.seek(0)
    return output


@report_bp.route('/export')
def export():
    data = report_data()
    f = data['filters']
    return send_file(make_workbook(data), as_attachment=True,
                     download_name=f"{data['title']}_{f['start_date']}_{f['end_date']}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
