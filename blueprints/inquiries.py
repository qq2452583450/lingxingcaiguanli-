"""
询价单蓝图
"""
from flask import Blueprint, request, jsonify, session, make_response
from datetime import datetime
from html import escape
from helpers import amount_to_chinese, get_db, generate_inquiry_no
from helpers.material_regions import (
    format_project_display,
    generate_material_code,
    get_region_name,
    resolve_material_region_code,
)
import sys
import logging
import config
sys.path.insert(0, '.')
from helpers.generate_inquiry_no import generate_inquiry_no_by_project

logger = logging.getLogger(__name__)

inquiry_bp = Blueprint('inquiries', __name__, url_prefix='/api')

SPECIAL_APPROVER_USERNAMES = ('leikefeng', 'tanxiang')


def _get_supplier_id_for_user(cursor, user_id):
    cursor.execute("SELECT id FROM suppliers WHERE user_id = ?", (user_id,))
    row = cursor.fetchone()
    return row['id'] if row else None


def _supplier_can_access_inquiry(cursor, inquiry_id, supplier_id):
    cursor.execute("""
        SELECT 1
        FROM purchase_inquiry_items pii
        JOIN purchase_inquiry_quotes piq ON piq.item_id = pii.id
        WHERE pii.inquiry_id = ?
          AND piq.supplier_id = ?
        LIMIT 1
    """, (inquiry_id, supplier_id))
    if cursor.fetchone():
        return True

    cursor.execute("""
        SELECT 1
        FROM purchase_inquiry_details
        WHERE inquiry_id = ?
          AND supplier_id = ?
        LIMIT 1
    """, (inquiry_id, supplier_id))
    return cursor.fetchone() is not None


def _get_special_approval_context(cursor, inquiry_id):
    cursor.execute("PRAGMA table_info(purchase_inquiries)")
    inquiry_columns = {row[1] for row in cursor.fetchall()}
    if 'project_id' in inquiry_columns:
        cursor.execute("""
            SELECT pi.id, pi.approval_status, pi.applicant_id,
                   p.project_code,
                   applicant.username AS applicant_username
            FROM purchase_inquiries pi
            LEFT JOIN projects p ON pi.project_id = p.id
            LEFT JOIN users applicant ON pi.applicant_id = applicant.id
            WHERE pi.id = ?
        """, (inquiry_id,))
    else:
        cursor.execute("""
            SELECT pi.id, pi.approval_status, pi.applicant_id,
                   NULL AS project_code,
                   applicant.username AS applicant_username
            FROM purchase_inquiries pi
            LEFT JOIN users applicant ON pi.applicant_id = applicant.id
            WHERE pi.id = ?
        """, (inquiry_id,))
    row = cursor.fetchone()
    if not row:
        return None, False
    ctx = dict(row)
    project_code = (ctx.get('project_code') or '').upper()
    applicant_username = (ctx.get('applicant_username') or '').lower()
    is_special = project_code.startswith('GX') or applicant_username == 'wanglihua'
    return ctx, is_special


def _get_recorded_special_approvers(cursor, inquiry_id):
    cursor.execute("""
        SELECT DISTINCT u.username
        FROM approval_records ar
        JOIN users u ON ar.approver_id = u.id
        WHERE ar.order_type = 'purchase_inquiry'
          AND ar.order_id = ?
          AND u.username IN (?, ?)
          AND (ar.result LIKE '%同意%' OR ar.result = '专项审批同意' OR ar.result = '主管同意')
    """, (inquiry_id, *SPECIAL_APPROVER_USERNAMES))
    return {row['username'] for row in cursor.fetchall()}


@inquiry_bp.route('/purchase-inquiries', methods=['GET'])
def get_inquiries():
    """获取询价单列表（根据用户权限过滤）"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    # 检查用户角色
    cursor.execute("SELECT r.role_name FROM users u LEFT JOIN roles r ON u.role_id = r.id WHERE u.id = ?", (user['id'],))
    role_row = cursor.fetchone()
    role_name = dict(role_row)['role_name'] if role_row else None
    cursor.execute("PRAGMA table_info(purchase_inquiries)")
    inquiry_columns = {row[1] for row in cursor.fetchall()}
    project_select = "p.project_code, p.project_name" if 'project_id' in inquiry_columns else "NULL AS project_code, NULL AS project_name"
    project_join = "LEFT JOIN projects p ON pi.project_id = p.id" if 'project_id' in inquiry_columns else ""

    keyword = (request.args.get('keyword') or '').strip()
    applicant = (request.args.get('applicant') or '').strip()
    status = (request.args.get('status') or '').strip()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    is_below = (request.args.get('is_below') or '').strip()

    filters = []
    params = []
    if keyword:
        filters.append("(pi.inquiry_no LIKE ? OR pi.remark LIKE ?)")
        like_keyword = f"%{keyword}%"
        params.extend([like_keyword, like_keyword])
    if applicant:
        filters.append("u.real_name LIKE ?")
        params.append(f"%{applicant}%")
    if status:
        filters.append("pi.approval_status = ?")
        params.append(status)
    if start_date:
        filters.append("pi.inquiry_date >= ?")
        params.append(start_date)
    if end_date:
        filters.append("pi.inquiry_date <= ?")
        params.append(end_date)
    if is_below in ('0', '1'):
        filters.append("pi.is_below_library_price = ?")
        params.append(int(is_below))

    if role_name == '系统管理员':
        # 管理员可以看到所有
        where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""
        cursor.execute(f"""
            SELECT pi.*, u.real_name as applicant_name, u.username as applicant_username,
                   {project_select}
            FROM purchase_inquiries pi
            LEFT JOIN users u ON pi.applicant_id = u.id
            {project_join}
            {where_sql}
            ORDER BY pi.create_time DESC
        """, params)
    elif role_name == '供应商':
        supplier_id = _get_supplier_id_for_user(cursor, user['id'])
        if not supplier_id:
            conn.close()
            return jsonify({'success': True, 'data': []})

        permission_sql = """(
                EXISTS (
                    SELECT 1
                    FROM purchase_inquiry_items pii
                    JOIN purchase_inquiry_quotes piq ON piq.item_id = pii.id
                    WHERE pii.inquiry_id = pi.id
                      AND piq.supplier_id = ?
                )
                OR EXISTS (
                    SELECT 1
                    FROM purchase_inquiry_details pid
                    WHERE pid.inquiry_id = pi.id
                      AND pid.supplier_id = ?
                )
            )"""
        where_parts = [permission_sql] + filters
        cursor.execute(f"""
            SELECT pi.*, u.real_name as applicant_name, u.username as applicant_username,
                   {project_select}
            FROM purchase_inquiries pi
            LEFT JOIN users u ON pi.applicant_id = u.id
            {project_join}
            WHERE {' AND '.join(where_parts)}
            ORDER BY pi.create_time DESC
        """, [supplier_id, supplier_id, *params])
    else:
        # 普通用户只能看到自己绑定项目的询价单
        permission_sql = """(
                pi.project_id IN (
                    SELECT project_id FROM user_projects WHERE user_id = ?
                ) OR pi.applicant_id = ?
            )"""
        where_parts = [permission_sql] + filters
        cursor.execute(f"""
            SELECT pi.*, u.real_name as applicant_name, u.username as applicant_username,
                   {project_select}
            FROM purchase_inquiries pi
            LEFT JOIN users u ON pi.applicant_id = u.id
            {project_join}
            WHERE {' AND '.join(where_parts)}
            ORDER BY pi.create_time DESC
        """, [user['id'], user['id'], *params])

    inquiries = []
    for row in cursor.fetchall():
        inquiry = dict(row)
        inquiry['project_city'] = get_region_name(inquiry.get('project_code'))
        inquiry['project_display_name'] = format_project_display(
            inquiry.get('project_code'),
            inquiry.get('project_name'),
        )
        inquiries.append(inquiry)
    conn.close()
    return jsonify({'success': True, 'data': inquiries})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>', methods=['GET'])
def get_inquiry(inquiry_id):
    """获取询价单详情（嵌套结构）"""
    user = session.get('user')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(purchase_inquiries)")
    inquiry_columns = {row[1] for row in cursor.fetchall()}
    project_select = "p.project_code, p.project_name" if 'project_id' in inquiry_columns else "NULL AS project_code, NULL AS project_name"
    project_join = "LEFT JOIN projects p ON pi.project_id = p.id" if 'project_id' in inquiry_columns else ""
    cursor.execute(f"""
        SELECT pi.*, u.real_name as applicant_name, u.username as applicant_username,
               {project_select}
        FROM purchase_inquiries pi
        LEFT JOIN users u ON pi.applicant_id = u.id
        {project_join}
        WHERE pi.id = ?
    """, (inquiry_id,))
    row = cursor.fetchone()
    inquiry = dict(row) if row else None
    supplier_id_for_response = None
    if inquiry:
        inquiry['project_city'] = get_region_name(inquiry.get('project_code'))
        inquiry['project_display_name'] = format_project_display(
            inquiry.get('project_code'),
            inquiry.get('project_name'),
        )

    if user:
        cursor.execute("SELECT r.role_name FROM users u LEFT JOIN roles r ON u.role_id = r.id WHERE u.id = ?", (user['id'],))
        role_row = cursor.fetchone()
        role_name = dict(role_row)['role_name'] if role_row else None
        if role_name == '供应商':
            supplier_id = _get_supplier_id_for_user(cursor, user['id'])
            if not supplier_id or not _supplier_can_access_inquiry(cursor, inquiry_id, supplier_id):
                conn.close()
                return jsonify({'success': False, 'message': '无权访问该询价单'}), 403
            supplier_id_for_response = supplier_id

    # 查询新的 items + quotes 嵌套结构
    item_supplier_filter = ""
    item_params = [inquiry_id]
    if supplier_id_for_response:
        item_supplier_filter = """
          AND EXISTS (
              SELECT 1
              FROM purchase_inquiry_quotes pq2
              WHERE pq2.item_id = pi.id
                AND pq2.supplier_id = ?
          )
        """
        item_params.append(supplier_id_for_response)
    cursor.execute(f"""
        SELECT pi.*, m.material_name, m.specification, m.material_code, u.unit_name
        FROM purchase_inquiry_items pi
        LEFT JOIN materials m ON pi.material_id = m.id
        LEFT JOIN units u ON m.unit_id = u.id
        WHERE pi.inquiry_id = ?
        {item_supplier_filter}
        ORDER BY pi.id
    """, item_params)
    items = []
    for item_row in cursor.fetchall():
        item = dict(item_row)
        item_id = item['id']
        # 查询该item下的所有报价
        quote_supplier_filter = ""
        quote_params = [item_id]
        if supplier_id_for_response:
            quote_supplier_filter = "AND pq.supplier_id = ?"
            quote_params.append(supplier_id_for_response)
        cursor.execute(f"""
            SELECT pq.*, s.supplier_name
            FROM purchase_inquiry_quotes pq
            LEFT JOIN suppliers s ON pq.supplier_id = s.id
            WHERE pq.item_id = ?
            {quote_supplier_filter}
            ORDER BY pq.id
        """, quote_params)
        item['quotes'] = [dict(q) for q in cursor.fetchall()]
        items.append(item)

    # 如果没有新结构数据，尝试兼容旧结构（purchase_inquiry_details）
    if not items:
        legacy_supplier_filter = ""
        legacy_params = [inquiry_id]
        if supplier_id_for_response:
            legacy_supplier_filter = "AND pd.supplier_id = ?"
            legacy_params.append(supplier_id_for_response)
        cursor.execute(f"""
            SELECT pd.*, m.material_name, m.specification, m.material_code,
                   u.unit_name, s.supplier_name
            FROM purchase_inquiry_details pd
            LEFT JOIN materials m ON pd.material_id = m.id
            LEFT JOIN units u ON m.unit_id = u.id
            LEFT JOIN suppliers s ON pd.supplier_id = s.id
            WHERE pd.inquiry_id = ?
            {legacy_supplier_filter}
        """, legacy_params)
        details = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return jsonify({'success': True, 'data': inquiry, 'details': details, 'legacy': True})

    conn.close()
    return jsonify({'success': True, 'data': inquiry, 'items': items})


@inquiry_bp.route('/purchase-inquiries', methods=['POST'])
def create_inquiry():
    """创建询价单（支持新的 items/quotes 嵌套结构）"""
    import json as json_module

    logger.info("=== create_inquiry called ===")

    # 手动解析 JSON
    try:
        data = json_module.loads(request.data)
    except Exception as e:
        logger.error("JSON parse error: %s", e)
        return jsonify({'success': False, 'message': 'JSON解析失败: ' + str(e)})

    if data is None:
        return jsonify({'success': False, 'message': '请求数据为空'})

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    # 判断数据格式：items 字段存在则走新结构，否则走旧结构兼容
    has_items_key = 'items' in data
    items_data = data.get('items', []) if has_items_key else []

    # 如果 items 字段不存在，尝试 details（旧结构兼容）
    if not has_items_key:
        items_data = data.get('details', [])

    is_new_format = has_items_key and isinstance(items_data, list)

    if is_new_format:
        # 新结构：items 格式
        if len(items_data) == 0:
            return jsonify({'success': False, 'message': '请添加询价材料（明细不能为空）'})

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 计算总金额（仅计算选定报价）
        total_amount = 0
        for item in items_data:
            selected_id = item.get('selected_quote_id')
            quantity = float(item.get('quantity', 1) or 1)
            for quote in item.get('quotes', []):
                tax_price = float(quote.get('tax_price', 0) or 0)
                # 只计入选定供应商的报价
                if selected_id and quote.get('supplier_id') == selected_id:
                    total_amount += tax_price * quantity
                # 如果没有选定，计入最低价
                elif not selected_id and quote.get('is_lowest'):
                    total_amount += tax_price * quantity

        # 判断是否低于库内价
        is_below = 0
        for item in items_data:
            library_price = float(item.get('library_price', 0) or 0)
            for quote in item.get('quotes', []):
                tax_price = float(quote.get('tax_price', 0) or 0)
                if library_price > 0 and tax_price < library_price:
                    is_below = 1
                    break
            if is_below:
                break

        # 写入主表
        inquiry_no = None
        inquiry_id = None
        
        # 检查表结构，确保 project_id 列存在
        cursor.execute("PRAGMA table_info(purchase_inquiries)")
        columns = [row[1] for row in cursor.fetchall()]
        has_project_id = 'project_id' in columns
        
        if not has_project_id:
            logger.warning("purchase_inquiries 表缺少 project_id 列，正在添加...")
            try:
                cursor.execute("ALTER TABLE purchase_inquiries ADD COLUMN project_id INTEGER")
                has_project_id = True
            except Exception as e:
                logger.error("添加 project_id 列失败: %s", e)
        
        for attempt in range(5):
            try:
                # 根据项目ID生成询价单号
                project_id = data.get('project_id')
                inquiry_date = data.get('inquiry_date', now[:10])
                if project_id:
                    inquiry_no = generate_inquiry_no_by_project(project_id, inquiry_date)
                else:
                    inquiry_no = generate_inquiry_no()

                # 根据表结构动态构建INSERT语句
                if has_project_id:
                    cursor.execute("""
                        INSERT INTO purchase_inquiries (
                            inquiry_no, inquiry_date, applicant_id, project_id, total_amount,
                            is_below_library_price, approval_status, create_time, remark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inquiry_no, inquiry_date,
                        user['id'], project_id, total_amount, is_below, '待审批', now, data.get('remark', '')
                    ))
                else:
                    cursor.execute("""
                        INSERT INTO purchase_inquiries (
                            inquiry_no, inquiry_date, applicant_id, total_amount,
                            is_below_library_price, approval_status, create_time, remark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inquiry_no, inquiry_date,
                        user['id'], total_amount, is_below, '待审批', now, data.get('remark', '')
                    ))
                inquiry_id = cursor.lastrowid
                break
            except Exception as insert_err:
                err_str = str(insert_err)
                if 'UNIQUE constraint failed' in err_str and attempt < 4:
                    logger.info("inquiry_no冲突，第%d次重试", attempt + 1)
                    conn.rollback()
                    continue
                conn.close()
                return jsonify({'success': False, 'message': f'创建失败: {err_str}'})

        if inquiry_id is None:
            conn.close()
            return jsonify({'success': False, 'message': '无法生成唯一单号，请稍后重试'})

        # 写入 items + quotes
        try:
            for item in items_data:
                material_id = item.get('material_id')
                quantity = float(item.get('quantity', 1) or 1)
                library_price = float(item.get('library_price', 0) or 0)
                selected_quote_id = item.get('selected_quote_id')
                tax_rate = float(item.get('tax_rate', 0.01) or 0.01)
                is_national_standard = item.get('is_national_standard', 0)
                is_cash_price = item.get('is_cash_price', 0)
                detail_spec = item.get('detail_spec', '') or '常规'
                brand = item.get('brand', '') or '无'

                # 如果选择了现金含税价，默认税率为1%
                if is_cash_price:
                    tax_rate = 0.01

                cursor.execute("""
                    INSERT INTO purchase_inquiry_items (
                        inquiry_id, material_id, quantity, library_price,
                        selected_quote_id, tax_rate, is_national_standard, is_cash_price,
                        detail_spec, brand, create_time
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    inquiry_id,
                    material_id if material_id else None,
                    quantity, library_price,
                    selected_quote_id, tax_rate, is_national_standard, is_cash_price,
                    detail_spec, brand, now
                ))
                item_id = cursor.lastrowid

                # 写入该item的所有报价（允许价格为0，留给供应商填写）
                quotes = item.get('quotes', [])
                valid_quotes = [q for q in quotes if q.get('supplier_id')]

                if valid_quotes:
                    # 计算最低价（基于不含税单价，仅考虑有价格的报价）
                    priced_quotes = [(float(q.get('tax_exempt_price', 0) or float(q.get('tax_price', 0)) / (1 + float(q.get('tax_rate', 0.13) or 0.13))), i) for i, q in enumerate(valid_quotes) if float(q.get('tax_price', 0) or 0) > 0]
                    lowest_idx = min(priced_quotes, key=lambda x: x[0])[1] if priced_quotes else -1

                    for i, quote in enumerate(valid_quotes):
                        tax_price = float(quote.get('tax_price', 0) or 0)
                        tax_exempt_price = float(quote.get('tax_exempt_price', 0) or 0)
                        tax_rate = float(quote.get('tax_rate', 0.13) or 0.13)
                        total = tax_price * quantity

                        # 自动计算不含税单价（如果未提供）
                        if tax_exempt_price == 0 and tax_price > 0:
                            tax_exempt_price = round(tax_price / (1 + tax_rate), 2)

                        # 有价格的报价为已提交，无价格的为待报价
                        q_status = 'submitted' if tax_price > 0 else 'pending'

                        cursor.execute("""
                            INSERT INTO purchase_inquiry_quotes (
                                item_id, supplier_id, tax_price, tax_exempt_price,
                                tax_rate, total_amount, is_lowest, is_selected, quote_status, create_time
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            item_id,
                            quote.get('supplier_id'),
                            tax_price, tax_exempt_price,
                            tax_rate, total,
                            1 if i == lowest_idx else 0,
                            1 if selected_quote_id and quote.get('supplier_id') == selected_quote_id else 0,
                            q_status, now
                        ))

            # 记录提交审批操作
            cursor.execute("""
                INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], '提交审批', '', now))

            conn.commit()
            conn.close()
            return jsonify({'success': True, 'inquiry_no': inquiry_no, 'id': inquiry_id})

        except Exception as e:
            logger.error("Error: %s", e)
            import traceback
            traceback.print_exc()
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)})

    else:
        # 旧结构兼容：直接使用 details
        details = data.get('details', [])
        if not isinstance(details, list):
            return jsonify({'success': False, 'message': '明细数据格式错误：应该为数组'})
        if len(details) == 0:
            return jsonify({'success': False, 'message': '明细数据不能为空'})

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        total_amount = sum(float(d.get('this_price', 0) or 0) * float(d.get('quantity', 1) or 1) for d in details)
        is_below = 1 if any(float(d.get('this_price', 0) or 0) < float(d.get('library_price', 0) or 0) for d in details) else 0

        inquiry_no = None
        inquiry_id = None
        
        # 检查表结构，确保 project_id 列存在
        cursor.execute("PRAGMA table_info(purchase_inquiries)")
        columns = [row[1] for row in cursor.fetchall()]
        has_project_id = 'project_id' in columns
        
        if not has_project_id:
            logger.warning("purchase_inquiries 表缺少 project_id 列，正在添加...")
            try:
                cursor.execute("ALTER TABLE purchase_inquiries ADD COLUMN project_id INTEGER")
                has_project_id = True
            except Exception as e:
                logger.error("添加 project_id 列失败: %s", e)

        for attempt in range(5):
            try:
                # 根据项目ID生成询价单号
                project_id = data.get('project_id')
                inquiry_date = data.get('inquiry_date', now[:10])
                if project_id:
                    inquiry_no = generate_inquiry_no_by_project(project_id, inquiry_date)
                else:
                    inquiry_no = generate_inquiry_no()

                # 根据表结构动态构建INSERT语句
                if has_project_id:
                    cursor.execute("""
                        INSERT INTO purchase_inquiries (
                            inquiry_no, inquiry_date, applicant_id, project_id, total_amount,
                            is_below_library_price, approval_status, create_time, remark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inquiry_no, inquiry_date,
                        user['id'], project_id, total_amount, is_below, '待审批', now, data.get('remark', '')
                    ))
                else:
                    cursor.execute("""
                        INSERT INTO purchase_inquiries (
                            inquiry_no, inquiry_date, applicant_id, total_amount,
                            is_below_library_price, approval_status, create_time, remark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        inquiry_no, inquiry_date,
                        user['id'], total_amount, is_below, '待审批', now, data.get('remark', '')
                    ))
                inquiry_id = cursor.lastrowid
                break
            except Exception as insert_err:
                err_str = str(insert_err)
                if 'UNIQUE constraint failed' in err_str and attempt < 4:
                    logger.info("inquiry_no冲突，第%d次重试", attempt + 1)
                    conn.rollback()
                    continue
                conn.close()
                return jsonify({'success': False, 'message': f'创建失败: {err_str}'})

        if inquiry_id is None:
            conn.close()
            return jsonify({'success': False, 'message': '无法生成唯一单号，请稍后重试'})

        try:
            for d in details:
                material_id = d.get('material_id')
                supplier_id = d.get('supplier_id')
                this_price = float(d.get('this_price', 0) or 0)
                library_price = float(d.get('library_price', 0) or 0)
                quantity = int(d.get('quantity', 1) or 1)
                tax_rate = float(d.get('tax_rate', 0.13) or 0.13)

                is_lowest = 1 if this_price <= library_price else 0
                price_diff = this_price - library_price

                cursor.execute("""
                    INSERT INTO purchase_inquiry_details (
                        inquiry_id, material_id, supplier_id, this_price,
                        library_price, is_lowest, price_diff, quantity, tax_rate
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    inquiry_id,
                    material_id if material_id else None,
                    supplier_id if supplier_id else None,
                    this_price, library_price,
                    is_lowest, price_diff, quantity, tax_rate
                ))

            # 记录提交审批操作
            cursor.execute("""
                INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], '提交审批', '', now))

            conn.commit()
            conn.close()
            return jsonify({'success': True, 'inquiry_no': inquiry_no, 'id': inquiry_id})

        except Exception as e:
            logger.error("Error: %s", e)
            import traceback
            traceback.print_exc()
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'message': str(e)})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>', methods=['DELETE'])
def delete_inquiry(inquiry_id):
    """删除询价单（含级联清理：入库单、库存、跨区域创建的材料）"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    # 获取询价单信息
    cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    inq = cursor.fetchone()
    if not inq:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})
    inq = dict(inq)
    inquiry_no = inq.get('inquiry_no', '')
    inquiry_project_id = inq.get('project_id')

    role_name = user.get('role_name')
    if not role_name:
        cursor.execute("SELECT r.role_name FROM users u LEFT JOIN roles r ON u.role_id = r.id WHERE u.id = ?", (user['id'],))
        role_row = cursor.fetchone()
        role_name = dict(role_row)['role_name'] if role_row else None

    can_delete = role_name == '系统管理员' or (
        role_name == '材料员' and inq.get('approval_status') == '草稿'
    )
    if not can_delete:
        conn.close()
        return jsonify({'success': False, 'message': '仅管理员或材料员可删除草稿询价单'})

    # 获取询价单所属项目的编码前缀
    inquiry_prefix = ''
    applicant_user = None
    cursor.execute("""
        SELECT username, real_name
        FROM users
        WHERE id = ?
    """, (inq.get('applicant_id'),))
    applicant_row = cursor.fetchone()
    if applicant_row:
        applicant_user = dict(applicant_row)

    if inquiry_project_id:
        cursor.execute("SELECT project_code FROM projects WHERE id = ?", (inquiry_project_id,))
        proj_row = cursor.fetchone()
        if proj_row:
            inquiry_prefix = resolve_material_region_code(proj_row[0], applicant_user)

    # 获取询价单关联的材料ID列表，识别跨区域创建的材料
    # 跨区域材料条件：(1) 编码前缀与项目前缀一致 (2) 仅被本询价单引用
    cursor.execute("SELECT material_id FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
    material_ids = [r[0] for r in cursor.fetchall()]

    cross_region_material_ids = []
    for mid in material_ids:
        # 检查该材料是否被其他询价单引用
        cursor.execute(
            "SELECT COUNT(*) FROM purchase_inquiry_items WHERE material_id = ? AND inquiry_id != ?",
            (mid, inquiry_id))
        if cursor.fetchone()[0] > 0:
            continue  # 材料被其他询价单共享，不删除

        # 检查该材料是否被入库明细引用
        cursor.execute("SELECT COUNT(*) FROM stock_in_details WHERE material_id = ?", (mid,))
        if cursor.fetchone()[0] > 0:
            continue  # 材料已入库，不删除

        # 检查该材料是否被采购订单引用
        cursor.execute("SELECT COUNT(*) FROM purchase_order_details WHERE material_id = ?", (mid,))
        if cursor.fetchone()[0] > 0:
            continue

        # 检查该材料是否被出库明细引用
        cursor.execute("SELECT COUNT(*) FROM stock_out_details WHERE material_id = ?", (mid,))
        if cursor.fetchone()[0] > 0:
            continue

        # 检查该材料是否被对账明细引用
        cursor.execute("SELECT COUNT(*) FROM reconciliation_details WHERE material_id = ?", (mid,))
        if cursor.fetchone()[0] > 0:
            continue

        # 检查该材料是否被基地库存引用
        cursor.execute("SELECT COUNT(*) FROM base_inventory WHERE material_id = ?", (mid,))
        if cursor.fetchone()[0] > 0:
            continue

        cursor.execute("SELECT material_code FROM materials WHERE id = ?", (mid,))
        mat_row = cursor.fetchone()
        if mat_row:
            mat_prefix = (mat_row[0] or '')[:2].upper()
            if mat_prefix == inquiry_prefix and inquiry_prefix:
                cross_region_material_ids.append(mid)

    try:
        # 1. 删除关联的入库单及明细，并扣减库存
        cursor.execute("SELECT id FROM stock_in_orders WHERE related_order_no = ?", (inquiry_no,))
        stock_in_ids = [r[0] for r in cursor.fetchall()]
        for si_id in stock_in_ids:
            cursor.execute("SELECT material_id, quantity, warehouse_id FROM stock_in_details WHERE order_id = ?", (si_id,))
            for detail in cursor.fetchall():
                mat_id = detail['material_id']
                qty = float(detail['quantity'] or 0)
                wh_id = detail['warehouse_id'] or 1
                cursor.execute("""
                    UPDATE inventory SET quantity = MAX(0, quantity - ?), update_time = ?
                    WHERE material_id = ? AND warehouse_id = ?
                """, (qty, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), mat_id, wh_id))
            cursor.execute("DELETE FROM stock_in_details WHERE order_id = ?", (si_id,))
        cursor.execute("DELETE FROM stock_in_orders WHERE related_order_no = ?", (inquiry_no,))

        # 2. 删除询价单明细和报价（含旧版 purchase_inquiry_details 表）
        cursor.execute("DELETE FROM purchase_inquiry_quotes WHERE item_id IN (SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?)", (inquiry_id,))
        cursor.execute("DELETE FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
        cursor.execute("DELETE FROM purchase_inquiry_details WHERE inquiry_id = ?", (inquiry_id,))
        cursor.execute("DELETE FROM approval_records WHERE order_type = 'purchase_inquiry' AND order_id = ?", (inquiry_id,))
        cursor.execute("DELETE FROM purchase_inquiries WHERE id = ?", (inquiry_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '删除成功'})

    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("删除询价单失败 inquiry_id=%s: %s", inquiry_id, e)
        return jsonify({'success': False, 'message': '删除失败: ' + str(e)})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>', methods=['PUT'])
def update_inquiry(inquiry_id):
    """编辑退回修改的询价单并重新提交"""
    import json as json_module

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    try:
        data = json_module.loads(request.data)
    except Exception as e:
        return jsonify({'success': False, 'message': 'JSON解析失败: ' + str(e)})

    conn = get_db()
    cursor = conn.cursor()

    # 校验：只有申请人可以编辑，且状态必须为退回修改
    cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    inquiry = dict(row)
    if inquiry['approval_status'] not in ('退回修改', '草稿'):
        conn.close()
        return jsonify({'success': False, 'message': '只有退回修改或草稿状态的询价单才能编辑'})
    if inquiry['applicant_id'] != user['id']:
        conn.close()
        return jsonify({'success': False, 'message': '只有申请人可以编辑此询价单'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 解析新明细数据
    items_data = data.get('items', [])
    if not items_data:
        conn.close()
        return jsonify({'success': False, 'message': '明细不能为空'})

    try:
        # 删除旧的 items 和 quotes
        cursor.execute("SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
        old_item_ids = [r[0] for r in cursor.fetchall()]
        if old_item_ids:
            placeholders = ','.join('?' * len(old_item_ids))
            cursor.execute(f"DELETE FROM purchase_inquiry_quotes WHERE item_id IN ({placeholders})", old_item_ids)
        cursor.execute("DELETE FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))

        # 重新计算总金额和是否低于库内价
        total_amount = 0
        is_below = 0
        for item in items_data:
            selected_id = item.get('selected_quote_id')
            quantity = float(item.get('quantity', 1) or 1)
            for quote in item.get('quotes', []):
                tax_price = float(quote.get('tax_price', 0) or 0)
                if selected_id and quote.get('supplier_id') == selected_id:
                    total_amount += tax_price * quantity
                elif not selected_id and quote.get('is_lowest'):
                    total_amount += tax_price * quantity
            library_price = float(item.get('library_price', 0) or 0)
            for quote in item.get('quotes', []):
                tp = float(quote.get('tax_price', 0) or 0)
                if library_price > 0 and tp < library_price:
                    is_below = 1
                    break

        # 写入新 items + quotes
        for item in items_data:
            material_id = item.get('material_id')
            quantity = float(item.get('quantity', 1) or 1)
            library_price = float(item.get('library_price', 0) or 0)
            selected_quote_id = item.get('selected_quote_id')
            tax_rate = float(item.get('tax_rate', 0.01) or 0.01)
            is_national_standard = item.get('is_national_standard', 0)
            is_cash_price = item.get('is_cash_price', 0)
            detail_spec = item.get('detail_spec', '') or '常规'
            brand = item.get('brand', '') or '无'

            if is_cash_price:
                tax_rate = 0.01

            cursor.execute("""
                INSERT INTO purchase_inquiry_items (
                    inquiry_id, material_id, quantity, library_price,
                    selected_quote_id, tax_rate, is_national_standard, is_cash_price,
                    detail_spec, brand, create_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                inquiry_id, material_id if material_id else None,
                quantity, library_price, selected_quote_id,
                tax_rate, is_national_standard, is_cash_price,
                detail_spec, brand, now
            ))
            item_id = cursor.lastrowid

            quotes = item.get('quotes', [])
            valid_quotes = [q for q in quotes if q.get('supplier_id')]

            if valid_quotes:
                # 计算最低价（仅考虑有价格的报价）
                priced_quotes = [(float(q.get('tax_exempt_price', 0) or float(q.get('tax_price', 0)) / (1 + float(q.get('tax_rate', 0.13) or 0.13))), i) for i, q in enumerate(valid_quotes) if float(q.get('tax_price', 0) or 0) > 0]
                lowest_idx = min(priced_quotes, key=lambda x: x[0])[1] if priced_quotes else -1

                for i, quote in enumerate(valid_quotes):
                    tp = float(quote.get('tax_price', 0) or 0)
                    tep = float(quote.get('tax_exempt_price', 0) or 0)
                    tr = float(quote.get('tax_rate', 0.13) or 0.13)
                    total = tp * quantity

                    if tep == 0 and tp > 0:
                        tep = round(tp / (1 + tr), 2)

                    q_status = 'submitted' if tp > 0 else 'pending'

                    cursor.execute("""
                        INSERT INTO purchase_inquiry_quotes (
                            item_id, supplier_id, tax_price, tax_exempt_price,
                            tax_rate, total_amount, is_lowest, is_selected, quote_status, create_time
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        item_id, quote.get('supplier_id'),
                        tp, tep, tr, total,
                        1 if i == lowest_idx else 0,
                        1 if selected_quote_id and quote.get('supplier_id') == selected_quote_id else 0,
                        q_status, now
                    ))

        # 更新主表并改为待审批
        cursor.execute("""
            UPDATE purchase_inquiries
            SET inquiry_date = ?, project_id = ?, total_amount = ?,
                is_below_library_price = ?, approval_status = '待审批',
                approver_id = NULL, approve_time = NULL, approval_remark = NULL,
                remark = ?
            WHERE id = ?
        """, (
            data.get('inquiry_date', inquiry['inquiry_date']),
            data.get('project_id', inquiry.get('project_id')),
            total_amount, is_below,
            data.get('remark', inquiry.get('remark', '')),
            inquiry_id
        ))

        # 记录重新提交审批
        cursor.execute("""
            INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], '重新提交', '', now))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '询价单已重新提交'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': str(e)})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/approve', methods=['POST'])
def approve_inquiry(inquiry_id):
    """审批询价单（支持新的 items/quotes 结构）"""
    try:
        return _approve_inquiry_impl(inquiry_id)
    except Exception as e:
        logger.error("审批异常: %s", e, exc_info=True)
        return jsonify({'success': False, 'message': f'审批失败: {str(e)}'})


def _approve_inquiry_impl(inquiry_id):
    """审批询价单的实际实现"""
    data = request.json
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    # 权限校验：只有系统管理员、材料员和材料审批负责人可以审批
    conn_check = get_db()
    cursor_check = conn_check.cursor()
    cursor_check.execute("SELECT r.role_name FROM users u LEFT JOIN roles r ON u.role_id = r.id WHERE u.id = ?", (user['id'],))
    role_row = cursor_check.fetchone()
    role_name = dict(role_row)['role_name'] if role_row else None
    if role_name not in ('系统管理员', '材料员', '材料审批负责人'):
        return jsonify({'success': False, 'message': '您没有审批权限，仅系统管理员、材料员或材料审批负责人可审批'})

    # 报价收集中时禁止审批，需先锁定报价
    cursor_check2 = get_db().cursor()
    cursor_check2.execute("SELECT quote_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    _qs = cursor_check2.fetchone()
    if _qs and dict(_qs).get('quote_status') == 'collecting':
        return jsonify({'success': False, 'message': '报价收集中，请先锁定报价后再审批'})

    action = data.get('action')
    remark = data.get('remark', '')

    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    special_ctx, requires_special_approval = _get_special_approval_context(cursor, inquiry_id)

    if requires_special_approval and role_name == '系统管理员':
        conn.close()
        return jsonify({'success': False, 'message': 'GX项目或王利华提交的询价单必须由雷克峰和谭香审批，admin不可审批'})

    if action == 'reject':
        cursor.execute("""
            UPDATE purchase_inquiries
            SET approval_status = '已驳回', approver_id = ?, approve_time = ?, approval_remark = ?
            WHERE id = ? AND approval_status IN ('待审批', '材料员已审', '报价未发布')
        """, (user['id'], now, remark, inquiry_id))
        updated_rows = cursor.rowcount
    elif action == 'return':
        # 退回：状态改回待修改，退回给申请人（支持已同意状态退回）
        # 先获取当前状态和询价单号
        cursor.execute("SELECT approval_status, inquiry_no FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
        inq_row = cursor.fetchone()
        if not inq_row:
            conn.close()
            return jsonify({'success': False, 'message': '询价单不存在'})
        current_status = inq_row['approval_status']
        inquiry_no = inq_row['inquiry_no']

        # 如果当前是已同意状态，需要反向处理自动入库的库存
        if current_status == '已同意':
            # 查找关联的入库单
            cursor.execute("""
                SELECT id, warehouse_id FROM stock_in_orders
                WHERE related_order_no = ? AND source_type = '采购入库' AND status = '已入库'
                ORDER BY create_time DESC LIMIT 1
            """, (inquiry_no,))
            stock_in = cursor.fetchone()
            if stock_in:
                stock_in_id = stock_in['id']
                warehouse_id = stock_in['warehouse_id'] or 1

                # 遍历入库明细，逐一扣减库存
                cursor.execute("SELECT * FROM stock_in_details WHERE order_id = ?", (stock_in_id,))
                for detail in cursor.fetchall():
                    material_id = detail['material_id']
                    detail_qty = float(detail['quantity'])

                    # 检查当前库存是否足够退回
                    cursor.execute("SELECT quantity FROM inventory WHERE material_id = ? AND warehouse_id = ?",
                                   (material_id, warehouse_id))
                    inv = cursor.fetchone()
                    current_qty = float(inv['quantity']) if inv else 0

                    if current_qty < detail_qty:
                        cursor.execute("SELECT material_name FROM materials WHERE id = ?", (material_id,))
                        mat = cursor.fetchone()
                        mat_name = mat['material_name'] if mat else f'ID:{material_id}'
                        conn.rollback()
                        conn.close()
                        return jsonify({
                            'success': False,
                            'message': f'无法退回：材料「{mat_name}」已有出库记录（库存 {current_qty} < 入库 {detail_qty}），请先处理出库单据后再退回'
                        })

                    # 扣减库存
                    cursor.execute("""
                        UPDATE inventory SET quantity = quantity - ?, update_time = ?
                        WHERE material_id = ? AND warehouse_id = ?
                    """, (detail_qty, now, material_id, warehouse_id))

                # 删除入库明细和入库单
                cursor.execute("DELETE FROM stock_in_details WHERE order_id = ?", (stock_in_id,))
                cursor.execute("DELETE FROM stock_in_orders WHERE id = ?", (stock_in_id,))

        # 更新询价单状态
        cursor.execute("""
            UPDATE purchase_inquiries
            SET approval_status = '退回修改', approver_id = ?, approve_time = ?, approval_remark = ?
            WHERE id = ? AND approval_status IN ('待审批', '材料员已审', '已同意', '报价未发布')
        """, (user['id'], now, remark, inquiry_id))
        updated_rows = cursor.rowcount
    elif action == 'material_clerk':
        cursor.execute("""
            UPDATE purchase_inquiries
            SET approval_status = '材料员已审', approver_id = ?, approve_time = ?, approval_remark = ?
            WHERE id = ? AND approval_status = '待审批'
        """, (user['id'], now, remark, inquiry_id))
        updated_rows = cursor.rowcount
    elif action == 'manager':
        if requires_special_approval:
            username = (user.get('username') or '').lower()
            if username not in SPECIAL_APPROVER_USERNAMES:
                conn.close()
                return jsonify({'success': False, 'message': '该询价单必须由雷克峰和谭香审批'})
            if special_ctx['approval_status'] not in ('待审批', '材料员已审', '退回修改', '报价未发布'):
                conn.close()
                return jsonify({'success': False, 'message': '操作失败，状态已更新'})

            approved_usernames = _get_recorded_special_approvers(cursor, inquiry_id)
            if username in approved_usernames:
                conn.close()
                return jsonify({'success': False, 'message': '您已审批过该询价单，需等待另一位负责人审批'})

            approved_after_current = approved_usernames | {username}
            if not all(name in approved_after_current for name in SPECIAL_APPROVER_USERNAMES):
                cursor.execute("""
                    INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], '专项审批同意', remark, now))
                conn.commit()
                conn.close()
                return jsonify({'success': True, 'message': '已记录专项审批，待另一位负责人审批后方可同意'})

        # 主管审批：待审批、材料员已审、退回修改 都可直接通过
        cursor.execute("""
            UPDATE purchase_inquiries
            SET approval_status = '已同意', approver_id = ?, approve_time = ?, approval_remark = ?
            WHERE id = ? AND approval_status IN ('待审批', '材料员已审', '退回修改', '报价未发布')
        """, (user['id'], now, remark, inquiry_id))
        updated_rows = cursor.rowcount

        # 尝试新的 items 结构
        cursor.execute("SELECT * FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
        items = cursor.fetchall()

        # 获取询价单所属项目的信息（用于生成新编号）
        cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
        inq = dict(cursor.fetchone())
        inquiry_project_id = inq.get('project_id')

        # 获取询价单所在项目的前缀（用于判断是否需要生成新编号）
        cursor.execute("SELECT project_code FROM projects WHERE id = ?", (inquiry_project_id,))
        proj_row = cursor.fetchone()
        inquiry_project_code = proj_row[0] if proj_row else ''
        cursor.execute("SELECT username, real_name FROM users WHERE id = ?", (inq.get('applicant_id'),))
        applicant_row = cursor.fetchone()
        inquiry_applicant = dict(applicant_row) if applicant_row else None
        inquiry_prefix = resolve_material_region_code(inquiry_project_code, inquiry_applicant) if inquiry_project_code else ''

        def generate_new_material_code(cursor, project_code):
            """根据项目编码生成新的材料编号（确保不重复）"""
            return generate_material_code(cursor, project_code, inquiry_applicant)

        if items:
            for item in items:
                item_dict = dict(item)
                material_id = item_dict['material_id']

                # 获取当前材料的编码前缀
                cursor.execute("SELECT material_code FROM materials WHERE id = ?", (material_id,))
                mat_row = cursor.fetchone()
                material_code = mat_row[0] if mat_row else ''
                material_prefix = material_code[:2].upper() if material_code and len(material_code) >= 2 else ''

                # 判断是否需要生成新编号（跨区域引用）
                need_new_code = (material_prefix != inquiry_prefix) and inquiry_prefix

                if need_new_code:
                    # 生成新编号
                    new_code = generate_new_material_code(cursor, inquiry_project_code)
                    logger.info("材料 %s 跨区域引用，创建新材料: %s -> %s", material_id, material_code, new_code)

                    # 查询原始材料完整信息
                    cursor.execute("SELECT * FROM materials WHERE id = ?", (material_id,))
                    orig_mat = dict(cursor.fetchone())

                    # 创建新材料（复制原始材料，更新编号和项目归属）
                    now_mat = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    # 优先使用询价单明细中用户填写的详细规格和品牌
                    new_brand = item_dict.get('brand') or orig_mat.get('brand', '')
                    new_detail_spec = item_dict.get('detail_spec') or orig_mat.get('detail_spec', '')
                    new_is_national_standard = item_dict.get('is_national_standard') if item_dict.get('is_national_standard') is not None else orig_mat.get('is_national_standard', 0)

                    cursor.execute("""
                        INSERT INTO materials (
                            material_code, material_name, specification, unit_id,
                            tax_price, tax_exempt_price, freight, remark,
                            default_supplier_id, inventory_min, inventory_max,
                            create_time, tax_rate, project_id, weight,
                            brand, is_national_standard, detail_spec,
                            cash_price, is_cash_price, cash_tax_price
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        new_code, orig_mat['material_name'], orig_mat['specification'],
                        orig_mat['unit_id'], orig_mat['tax_price'], orig_mat['tax_exempt_price'],
                        orig_mat['freight'], orig_mat['remark'], orig_mat['default_supplier_id'],
                        orig_mat['inventory_min'], orig_mat['inventory_max'], now_mat,
                        orig_mat['tax_rate'], inquiry_project_id, orig_mat.get('weight', 0),
                        new_brand, new_is_national_standard, new_detail_spec,
                        orig_mat['cash_price'], orig_mat['is_cash_price'], orig_mat['cash_tax_price']
                    ))
                    new_material_id = cursor.lastrowid

                    # 更新询价单明细，指向新材料
                    cursor.execute("UPDATE purchase_inquiry_items SET material_id = ? WHERE id = ?",
                                   (new_material_id, item_dict['id']))
                    material_id = new_material_id
                    logger.info("已创建新材料 id=%s，询价明细 item_id=%s 已更新关联", new_material_id, item_dict['id'])

                # 查找选定的报价（is_selected=1）
                cursor.execute("""
                    SELECT * FROM purchase_inquiry_quotes
                    WHERE item_id = ? AND is_selected = 1 LIMIT 1
                """, (item_dict['id'],))
                selected_quote = cursor.fetchone()

                if selected_quote:
                    selected_quote = dict(selected_quote)
                    quote_tax_rate = float(selected_quote.get('tax_rate', 0.01) or 0.01)
                    is_cash = item_dict.get('is_cash_price', 0)
                    # 根据选定的报价更新材料价格（need_new_code时material_id已指向新材料）
                    if is_cash:
                        cash_tax = round(selected_quote['tax_price'] / (1 + quote_tax_rate), 2) if quote_tax_rate else selected_quote['tax_price']
                        cursor.execute(
                            "UPDATE materials SET cash_price = ?, cash_tax_price = ?, default_supplier_id = ?, tax_rate = ?, is_cash_price = 1 WHERE id = ?",
                            (selected_quote['tax_price'], cash_tax, selected_quote['supplier_id'], quote_tax_rate, material_id)
                        )
                    else:
                        cursor.execute(
                            "UPDATE materials SET tax_price = ?, tax_exempt_price = ?, default_supplier_id = ?, tax_rate = ? WHERE id = ?",
                            (selected_quote['tax_price'], selected_quote['tax_exempt_price'], selected_quote['supplier_id'], quote_tax_rate, material_id)
                        )
                else:
                    # 如果没有选定，使用最低价报价
                    cursor.execute("""
                        SELECT * FROM purchase_inquiry_quotes
                        WHERE item_id = ? AND is_lowest = 1 LIMIT 1
                    """, (item_dict['id'],))
                    lowest_quote = cursor.fetchone()
                    if lowest_quote:
                        lowest_quote = dict(lowest_quote)
                        quote_tax_rate = float(lowest_quote.get('tax_rate', 0.01) or 0.01)
                        is_cash = item_dict.get('is_cash_price', 0)
                        if is_cash:
                            cash_tax = round(lowest_quote['tax_price'] / (1 + quote_tax_rate), 2) if quote_tax_rate else lowest_quote['tax_price']
                            cursor.execute(
                                "UPDATE materials SET cash_price = ?, cash_tax_price = ?, default_supplier_id = ?, tax_rate = ?, is_cash_price = 1 WHERE id = ?",
                                (lowest_quote['tax_price'], cash_tax, lowest_quote['supplier_id'], quote_tax_rate, material_id)
                            )
                        else:
                            cursor.execute(
                                "UPDATE materials SET tax_price = ?, tax_exempt_price = ?, default_supplier_id = ?, tax_rate = ? WHERE id = ?",
                                (lowest_quote['tax_price'], lowest_quote['tax_exempt_price'], lowest_quote['supplier_id'], quote_tax_rate, material_id)
                            )
        else:
            # 兼容旧结构：purchase_inquiry_details
            cursor.execute("SELECT * FROM purchase_inquiry_details WHERE inquiry_id = ?", (inquiry_id,))
            for d in cursor.fetchall():
                material_id = d['material_id']
                tax_price = d['this_price']
                tax_exempt = round(tax_price / 1.01, 2)
                supplier_id = d['supplier_id']

                # 获取当前材料的编码前缀
                cursor.execute("SELECT material_code FROM materials WHERE id = ?", (material_id,))
                mat_row = cursor.fetchone()
                material_code = mat_row[0] if mat_row else ''
                material_prefix = material_code[:2].upper() if material_code and len(material_code) >= 2 else ''

                # 判断是否需要生成新编号（跨区域引用）
                need_new_code = (material_prefix != inquiry_prefix) and inquiry_prefix

                if need_new_code:
                    # 生成新编号
                    new_code = generate_new_material_code(cursor, inquiry_project_code)
                    logger.info("材料 %s 跨区域引用，创建新材料: %s -> %s", material_id, material_code, new_code)

                    # 查询原始材料完整信息
                    cursor.execute("SELECT * FROM materials WHERE id = ?", (material_id,))
                    orig_mat = dict(cursor.fetchone())

                    # 创建新材料（复制原始材料，更新编号和项目归属）
                    now_mat = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute("""
                        INSERT INTO materials (
                            material_code, material_name, specification, unit_id,
                            tax_price, tax_exempt_price, freight, remark,
                            default_supplier_id, inventory_min, inventory_max,
                            create_time, tax_rate, project_id, weight,
                            brand, is_national_standard, detail_spec,
                            cash_price, is_cash_price, cash_tax_price
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        new_code, orig_mat['material_name'], orig_mat['specification'],
                        orig_mat['unit_id'], orig_mat['tax_price'], orig_mat['tax_exempt_price'],
                        orig_mat['freight'], orig_mat['remark'], orig_mat['default_supplier_id'],
                        orig_mat['inventory_min'], orig_mat['inventory_max'], now_mat,
                        orig_mat['tax_rate'], inquiry_project_id, orig_mat.get('weight', 0),
                        orig_mat['brand'], orig_mat['is_national_standard'], orig_mat['detail_spec'],
                        orig_mat['cash_price'], orig_mat['is_cash_price'], orig_mat['cash_tax_price']
                    ))
                    new_material_id = cursor.lastrowid

                    # 更新询价单明细，指向新材料
                    cursor.execute("UPDATE purchase_inquiry_details SET material_id = ? WHERE id = ?",
                                   (new_material_id, d['id']))
                    material_id = new_material_id
                    logger.info("已创建新材料 id=%s，询价明细已更新关联", new_material_id)

                # 更新价格
                if supplier_id:
                    cursor.execute(
                        "UPDATE materials SET tax_price = ?, tax_exempt_price = ?, default_supplier_id = ? WHERE id = ?",
                        (tax_price, tax_exempt, supplier_id, material_id)
                    )
                else:
                    cursor.execute(
                        "UPDATE materials SET tax_price = ?, tax_exempt_price = ? WHERE id = ?",
                        (tax_price, tax_exempt, material_id)
                    )

        cursor.execute("UPDATE purchase_inquiries SET library_price_updated = 1 WHERE id = ?", (inquiry_id,))

        # 获取询价单信息
        cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
        inq = dict(cursor.fetchone())
        inquiry_no = inq.get('inquiry_no', '')

        # ===== 审批通过后自动入库 =====
        # 安全防护：删除旧的活跃入库单（防止重复入库）
        cursor.execute("""
            SELECT id, warehouse_id FROM stock_in_orders
            WHERE related_order_no = ? AND source_type = '采购入库' AND status = '已入库'
        """, (inquiry_no,))
        old_stock_ins = cursor.fetchall()
        for old_si in old_stock_ins:
            old_si_id = old_si['id']
            old_warehouse_id = old_si['warehouse_id'] or 1
            # 反向扣减旧入库单的库存（使用 MAX 防止负库存）
            cursor.execute("SELECT * FROM stock_in_details WHERE order_id = ?", (old_si_id,))
            for old_detail in cursor.fetchall():
                old_material_id = old_detail['material_id']
                old_qty = float(old_detail['quantity'])
                cursor.execute("""
                    UPDATE inventory SET quantity = MAX(0, quantity - ?), update_time = ?
                    WHERE material_id = ? AND warehouse_id = ?
                """, (old_qty, now, old_material_id, old_warehouse_id))
            # 删除旧入库单及明细
            cursor.execute("DELETE FROM stock_in_details WHERE order_id = ?", (old_si_id,))
            cursor.execute("DELETE FROM stock_in_orders WHERE id = ?", (old_si_id,))

        # 生成入库单号（取当日最大编号+1，避免删除后COUNT冲突）
        today_str = datetime.now().strftime('%Y%m%d')
        cursor.execute(
            "SELECT order_no FROM stock_in_orders WHERE order_no LIKE ? ORDER BY order_no DESC LIMIT 1",
            (f'JH-{today_str}%',)
        )
        last_row = cursor.fetchone()
        if last_row:
            try:
                last_num = int(last_row[0].split('-')[-1])
            except (ValueError, IndexError):
                last_num = 0
            stk_count = last_num + 1
        else:
            stk_count = 1
        stock_in_no = f'JH-{today_str}-{str(stk_count).zfill(3)}'

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 创建入库主单
        cursor.execute("""
            INSERT INTO stock_in_orders (
                order_no, source_type, related_order_no, supplier_id,
                warehouse_id, project_id, operator_id, in_time, status, create_time, remark
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            stock_in_no, '采购入库', inquiry_no, None,
            1, inq.get('project_id'), user['id'], now_str, '已入库', now_str,
            f'询价单{inq.get("inquiry_no", "")}审批通过自动入库'
        ))
        stock_in_id = cursor.lastrowid

        # 遍历选定报价，创建入库明细并更新库存
        if items:
            for item in items:
                item_dict = dict(item)
                # 重新查询material_id（前面审批可能已创建新材料并更新了关联）
                cursor.execute("SELECT material_id FROM purchase_inquiry_items WHERE id = ?", (item_dict['id'],))
                mid_row = cursor.fetchone()
                material_id = mid_row[0] if mid_row else item_dict['material_id']
                quantity = item_dict.get('quantity', 1)

                unit_price = 0
                supplier_id = None

                # 优先查找 is_selected=1 的报价（前端选定的拟定供应商）
                cursor.execute("""
                    SELECT * FROM purchase_inquiry_quotes
                    WHERE item_id = ? AND is_selected = 1 LIMIT 1
                """, (item_dict['id'],))
                q = cursor.fetchone()

                # 如果没有 is_selected 的，退而查找 is_lowest=1 的最低价报价
                if not q:
                    cursor.execute("""
                        SELECT * FROM purchase_inquiry_quotes
                        WHERE item_id = ? AND is_lowest = 1 LIMIT 1
                    """, (item_dict['id'],))
                    q = cursor.fetchone()

                # 最后兜底：取该 item 下任意一条有效报价
                if not q:
                    cursor.execute("""
                        SELECT * FROM purchase_inquiry_quotes
                        WHERE item_id = ? AND tax_price > 0 LIMIT 1
                    """, (item_dict['id'],))
                    q = cursor.fetchone()

                if q:
                    unit_price = q['tax_price'] or 0
                    supplier_id = q['supplier_id']

                if unit_price > 0 and quantity > 0:
                    amount = unit_price * quantity
                    # 入库明细（记录每个材料对应的供应商）
                    cursor.execute("""
                        INSERT INTO stock_in_details (order_id, material_id, quantity, unit_price, amount, supplier_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (stock_in_id, material_id, quantity, unit_price, amount, supplier_id))

                    # 更新库存
                    cursor.execute("""
                        SELECT id FROM inventory WHERE material_id = ? AND warehouse_id = ?
                    """, (material_id, 1))
                    existing_inv = cursor.fetchone()
                    if existing_inv:
                        cursor.execute("""
                            UPDATE inventory SET quantity = quantity + ?, unit_price = ?, update_time = ?
                            WHERE material_id = ? AND warehouse_id = ?
                        """, (quantity, unit_price, now_str, material_id, 1))
                    else:
                        cursor.execute("""
                            INSERT INTO inventory (material_id, warehouse_id, quantity, unit_price, update_time)
                            VALUES (?, ?, ?, ?, ?)
                        """, (material_id, 1, quantity, unit_price, now_str))

                    # 更新入库单的供应商（用第一个有供应商的报价）
                    if supplier_id:
                        cursor.execute("UPDATE stock_in_orders SET supplier_id = ? WHERE id = ? AND supplier_id IS NULL", (supplier_id, stock_in_id))
        else:
            # 兼容旧结构
            cursor.execute("SELECT * FROM purchase_inquiry_details WHERE inquiry_id = ?", (inquiry_id,))
            for d in cursor.fetchall():
                d_dict = dict(d)
                material_id = d_dict['material_id']
                quantity = d_dict.get('quantity', 1)
                unit_price = d_dict.get('this_price', 0)
                supplier_id = d_dict.get('supplier_id')

                if unit_price > 0 and quantity > 0:
                    amount = unit_price * quantity
                    cursor.execute("""
                        INSERT INTO stock_in_details (order_id, material_id, quantity, unit_price, amount)
                        VALUES (?, ?, ?, ?, ?)
                    """, (stock_in_id, material_id, quantity, unit_price, amount))

                    cursor.execute("""
                        SELECT id FROM inventory WHERE material_id = ? AND warehouse_id = ?
                    """, (material_id, 1))
                    existing_inv = cursor.fetchone()
                    if existing_inv:
                        cursor.execute("""
                            UPDATE inventory SET quantity = quantity + ?, unit_price = ?, update_time = ?
                            WHERE material_id = ? AND warehouse_id = ?
                        """, (quantity, unit_price, now_str, material_id, 1))
                    else:
                        cursor.execute("""
                            INSERT INTO inventory (material_id, warehouse_id, quantity, unit_price, update_time)
                            VALUES (?, ?, ?, ?, ?)
                        """, (material_id, 1, quantity, unit_price, now_str))

                    if supplier_id:
                        cursor.execute("UPDATE stock_in_orders SET supplier_id = ? WHERE id = ? AND supplier_id IS NULL", (supplier_id, stock_in_id))
        # ===== 自动入库结束 =====
    else:
        conn.close()
        return jsonify({'success': False, 'message': '无效的操作'})

    if updated_rows == 0:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': '操作失败，状态已更新'})

    result_text = {'reject': '驳回', 'return': '退回', 'material_clerk': '材料员同意', 'manager': '主管同意'}.get(action, action)
    cursor.execute("""
        INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], result_text, remark, now))

    conn.commit()
    conn.close()
    return jsonify({'success': True})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/recall', methods=['POST'])
def recall_inquiry(inquiry_id):
    """撤回询价单（仅申请人在待审批状态下可撤回，撤回后状态变为草稿）"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    # 查询询价单
    cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    inquiry = dict(row)

    # 校验：只有申请人可以撤回
    if inquiry['applicant_id'] != user['id']:
        conn.close()
        return jsonify({'success': False, 'message': '只有申请人可以撤回此询价单'})

    # 校验：只有待审批状态可以撤回
    if inquiry['approval_status'] != '待审批':
        conn.close()
        return jsonify({'success': False, 'message': '只有待审批状态的询价单才能撤回'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 更新状态为草稿，清除审批信息
    cursor.execute("""
        UPDATE purchase_inquiries
        SET approval_status = '草稿', approver_id = NULL, approve_time = NULL, approval_remark = NULL
        WHERE id = ? AND approval_status = '待审批'
    """, (inquiry_id,))
    updated_rows = cursor.rowcount

    if updated_rows == 0:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': '操作失败，状态已更新'})

    # 记录撤回操作到审批记录
    cursor.execute("""
        INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, ('purchase_inquiry', inquiry_id, user['id'], user['real_name'], '撤回', '', now))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '询价单已撤回'})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/approval-history', methods=['GET'])
def get_inquiry_approval_history(inquiry_id):
    """获取询价单审批历史"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT ar.*, u.real_name as approver_real_name
        FROM approval_records ar
        LEFT JOIN users u ON ar.approver_id = u.id
        WHERE ar.order_type = 'purchase_inquiry' AND ar.order_id = ?
        ORDER BY ar.approval_time ASC
    """, (inquiry_id,))
    records = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': records})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/approval-print', methods=['GET'])
def print_inquiry_approval(inquiry_id):
    """打印询价单审批签字单（支持新的 items/quotes 嵌套结构）"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT pi.*, u.real_name as applicant_name, p.project_name
        FROM purchase_inquiries pi
        LEFT JOIN users u ON pi.applicant_id = u.id
        LEFT JOIN projects p ON pi.project_id = p.id
        WHERE pi.id = ?
    """, (inquiry_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '单据不存在'})

    inquiry = dict(row)

    # 尝试新的 items + quotes 结构
    cursor.execute("""
        SELECT pi.*, m.material_name, m.specification, m.material_code, u.unit_name
        FROM purchase_inquiry_items pi
        LEFT JOIN materials m ON pi.material_id = m.id
        LEFT JOIN units u ON m.unit_id = u.id
        WHERE pi.inquiry_id = ?
        ORDER BY pi.id
    """, (inquiry_id,))
    items = []
    for item_row in cursor.fetchall():
        item = dict(item_row)
        item_id = item['id']
        cursor.execute("""
            SELECT pq.*, s.supplier_name
            FROM purchase_inquiry_quotes pq
            LEFT JOIN suppliers s ON pq.supplier_id = s.id
            WHERE pq.item_id = ?
            ORDER BY pq.id
        """, (item_id,))
        item['quotes'] = [dict(q) for q in cursor.fetchall()]
        items.append(item)

    # 如果没有新结构数据，使用旧结构
    if not items:
        cursor.execute("""
            SELECT pd.*, m.material_name, m.specification, m.material_code, u.unit_name, s.supplier_name
            FROM purchase_inquiry_details pd
            LEFT JOIN materials m ON pd.material_id = m.id
            LEFT JOIN units u ON m.unit_id = u.id
            LEFT JOIN suppliers s ON pd.supplier_id = s.id
            WHERE pd.inquiry_id = ?
        """, (inquiry_id,))
        details = [dict(row) for row in cursor.fetchall()]
    else:
        details = None

    cursor.execute("""
        SELECT ar.*, u.real_name as approver_real_name
        FROM approval_records ar
        LEFT JOIN users u ON ar.approver_id = u.id
        WHERE ar.order_type = 'purchase_inquiry' AND ar.order_id = ?
        ORDER BY ar.approval_time ASC
    """, (inquiry_id,))
    approval_records = [dict(row) for row in cursor.fetchall()]

    conn.close()

    amount_chinese = amount_to_chinese(inquiry['total_amount'])

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>询价单审批签字单 - {inquiry['inquiry_no']}</title>
        <style>
            @page {{ size: A4 landscape; margin: 5mm; }}
            body {{ font-family: "微软雅黑", "Microsoft YaHei", Arial, sans-serif; margin: 0; padding: 10px; box-sizing: border-box; }}
            .print-page {{ width: 100%; max-width: 287mm; margin: 0 auto; background: white; box-sizing: border-box; overflow: visible; }}
            .print-header {{ text-align: center; margin-bottom: 8px; border-bottom: 2px solid #333; padding-bottom: 6px; }}
            .print-header h1 {{ margin: 0 0 3px 0; font-size: 18px; }}
            .print-header div {{ font-size: 12px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 5px 0; table-layout: fixed; }}
            th, td {{ border: 1px solid #333; padding: 3px 4px; font-size: 9px; word-wrap: break-word; overflow-wrap: break-word; line-height: 1.2; }}
            th {{ background: #f0f0f0; font-size: 9px; }}
            .col-seq {{ width: 3%; }}
            .col-name {{ width: 9%; }}
            .col-spec {{ width: 6%; }}
            .col-detail {{ width: 8%; }}
            .col-unit {{ width: 4%; }}
            .col-qty {{ width: 4%; }}
            .col-lib {{ width: 5%; }}
            .col-supplier {{ width: 14%; }}
            .col-price {{ width: 7%; }}
            .col-tax {{ width: 4%; }}
            .col-tax-exempt {{ width: 7%; }}
            .col-amount {{ width: 7%; }}
            .col-lowest {{ width: 4%; }}
            .col-selected {{ width: 4%; }}
            .info-row {{ display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 5px; font-size: 11px; }}
            .info-item {{ flex: 1; min-width: 180px; }}
            .info-item label {{ font-weight: bold; font-size: 11px; }}
            .approval-section {{ margin-top: 8px; }}
            .approval-title {{ font-size: 12px; font-weight: bold; margin-bottom: 5px; border-bottom: 1px solid #333; padding-bottom: 3px; }}
            .approval-table {{ margin-bottom: 8px; }}
            .approval-table th, .approval-table td {{ font-size: 9px; padding: 2px 4px; }}
            .signatures {{ display: flex; justify-content: space-between; margin-top: 10px; }}
            .signature-item {{ text-align: center; width: 20%; }}
            .signature-line {{ border-top: 1px solid #333; margin-top: 10px; padding-top: 3px; font-size: 10px; }}
            .amount-chinese {{ font-size: 12px; font-weight: bold; color: #c0392b; margin-top: 5px; }}
            .total-section {{ clear: both; text-align: right; margin: 5px 0; padding: 5px; background: #f9f9f9; border: 1px solid #ddd; font-size: 11px; }}
            .item-card {{ border: 1px solid #ddd; margin-bottom: 15px; border-radius: 4px; }}
            .item-header {{ background: #f5f5f5; padding: 8px 10px; border-bottom: 1px solid #ddd; font-weight: bold; }}
            .quote-row {{ display: flex; padding: 6px 10px; border-bottom: 1px solid #eee; }}
            .quote-row:last-child {{ border-bottom: none; }}
            .quote-row.lowest {{ background: #f0fff4; }}
            .quote-supplier {{ width: 150px; }}
            .quote-price {{ width: 100px; text-align: right; }}
            .quote-amount {{ width: 100px; text-align: right; font-weight: bold; }}
            .quote-selected {{ width: 60px; color: #27ae60; font-weight: bold; }}
            .lowest-tag {{ color: #27ae60; font-size: 9px; }}
            .print-btn {{ display: inline-block; padding: 8px 25px; background: #3498db; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; margin: 10px auto; }}
            .print-btn:hover {{ background: #2980b9; }}
            @media print {{
                .no-print {{ display: none !important; }}
                body {{ padding: 0; margin: 0; }}
                .print-page {{ max-width: 100%; padding: 5px; }}
                table {{ font-size: 9px; margin: 3px 0; }}
                th, td {{ padding: 2px 3px; }}
                .approval-section {{ margin-top: 5px; }}
                .approval-title {{ margin-bottom: 3px; }}
                .approval-table th, .approval-table td {{ padding: 2px 3px; }}
                .signatures {{ margin-top: 8px; }}
                .signature-line {{ margin-top: 8px; padding-top: 2px; }}
                .total-section {{ margin: 3px 0; padding: 3px; }}
                .info-row {{ margin-bottom: 3px; gap: 3px; }}
            }}
        </style>
    </head>
    <body>
        <div class="no-print" style="text-align:right;margin:10px 20px;">
            <button class="print-btn" onclick="window.print()">打印签字单</button>
        </div>
        <div class="print-page">
            <div class="print-header">
                <h1>采购询比价审批签字单</h1>
                <div>{escape(inquiry['inquiry_no'])}</div>
            </div>

            <div class="info-row">
                <div class="info-item"><label>申请日期：</label>{escape(inquiry.get('inquiry_date', '-'))}</div>
                <div class="info-item"><label>申请人：</label>{escape(inquiry.get('applicant_name', '-'))}</div>
                <div class="info-item"><label>低于库内价：</label>{'是' if inquiry.get('is_below_library_price') == 1 else '否'}</div>
                <div class="info-item"><label>项目：</label>{escape(inquiry.get('project_name') or '-')}</div>
            </div>
            <div class="info-row">
                <div class="info-item"><label>当前状态：</label><strong>{escape(inquiry.get('approval_status', '-'))}</strong></div>
            </div>
"""

    if items:
        # 新结构：嵌套展示，材料名+规格相同行合并（rowspan）
        html += """<table>
                <thead>
                    <tr>
                        <th class="col-seq">序号</th>
                        <th class="col-name">名称</th>
                        <th class="col-spec">规格</th>
                        <th class="col-detail">详细规格</th>
                        <th class="col-unit">单位</th>
                        <th class="col-qty">数量</th>
                        <th class="col-lib">库内价</th>
                        <th class="col-supplier">供应商</th>
                        <th class="col-price">含税单价</th>
                        <th class="col-tax">税率</th>
                        <th class="col-tax-exempt">不含税单价</th>
                        <th class="col-amount">报价金额</th>
                        <th class="col-lowest">最低价</th>
                        <th class="col-selected">拟定</th>
                    </tr>
                </thead>
                <tbody>
"""
        for idx, item in enumerate(items):
            material_name = item.get('material_name', '')
            specification = item.get('specification', '-')
            detail_spec = item.get('detail_spec', '') or '常规'
            unit_name = item.get('unit_name', '-')
            quantity = item.get('quantity', 1)
            library_price = item.get('library_price', 0)
            quotes = item.get('quotes', [])

            if not quotes:
                # 无报价时单行显示
                html += f"""
                    <tr>
                        <td>{idx + 1}</td>
                        <td style="font-weight:500;">{escape(material_name)}</td>
                        <td>{escape(specification)}</td>
                        <td>{escape(detail_spec)}</td>
                        <td>{escape(unit_name)}</td>
                        <td>{quantity}</td>
                        <td>¥{library_price:.2f}</td>
                        <td>-</td>
                        <td>-</td>
                        <td>-</td>
                        <td>-</td>
                        <td>-</td>
                        <td></td>
                        <td></td>
                    </tr>
"""
            else:
                rowspan = f' rowspan="{len(quotes)}"' if len(quotes) > 1 else ''
                for qi, quote in enumerate(quotes):
                    tax_rate = float(quote.get('tax_rate', 0.01) or 0.01)
                    html += "<tr>"
                    if qi == 0:
                        html += f'<td{rowspan} style="vertical-align:middle;">{idx + 1}</td>'
                        html += f'<td{rowspan} style="font-weight:500;vertical-align:middle;">{escape(material_name)}</td>'
                        html += f'<td{rowspan} style="vertical-align:middle;">{escape(specification)}</td>'
                        html += f'<td{rowspan} style="vertical-align:middle;">{escape(detail_spec)}</td>'
                        html += f'<td{rowspan} style="vertical-align:middle;">{escape(unit_name)}</td>'
                        html += f'<td{rowspan} style="vertical-align:middle;">{quantity}</td>'
                        html += f'<td{rowspan} style="vertical-align:middle;">¥{library_price:.2f}</td>'
                    tax_price = float(quote.get('tax_price', 0) or 0)
                    tax_exempt_price = round(tax_price / (1 + tax_rate), 2) if tax_rate > -1 else tax_price
                    html += f'<td>{escape(quote.get("supplier_name", "-"))}</td>'
                    html += f'<td>¥{tax_price:.2f}</td>'
                    html += f'<td>{tax_rate * 100:.0f}%</td>'
                    html += f'<td>¥{tax_exempt_price:.2f}</td>'
                    html += f'<td>¥{quote.get("total_amount", 0):.2f}</td>'
                    lowest_html = '<span class="lowest-tag">最低</span>' if quote.get('is_lowest') == 1 else ''
                    html += f'<td>{lowest_html}</td>'
                    selected_html = '✓' if quote.get('is_selected') == 1 else ''
                    html += f'<td>{selected_html}</td>'
                    html += "</tr>\n"
        html += """                </tbody>
            </table>
"""
    else:
        # 旧结构：扁平展示，材料名+规格相同行合并（rowspan）
        html += """<table>
                <thead>
                    <tr>
                        <th class="col-seq">序号</th>
                        <th class="col-name">名称</th>
                        <th class="col-spec">规格</th>
                        <th class="col-detail">详细规格</th>
                        <th class="col-unit">单位</th>
                        <th class="col-lib">库内价</th>
                        <th class="col-supplier">供应商</th>
                        <th class="col-price">本次报价</th>
                        <th class="col-amount">差额</th>
                    </tr>
                </thead>
                <tbody>
"""
        # 按材料名+规格分组
        from itertools import groupby
        def detail_group_key(d):
            return (d.get('material_name', ''), d.get('specification', '-'), d.get('detail_spec', ''))
        sorted_details = sorted(details, key=detail_group_key)
        for gidx, (key, group_iter) in enumerate(groupby(sorted_details, key=detail_group_key)):
            group_rows = list(group_iter)
            rowspan = f' rowspan="{len(group_rows)}"' if len(group_rows) > 1 else ''
            for gi, d in enumerate(group_rows):
                html += "<tr>"
                if gi == 0:
                    html += f'<td{rowspan} style="vertical-align:middle;">{gidx + 1}</td>'
                    html += f'<td{rowspan} style="font-weight:500;vertical-align:middle;">{escape(key[0])}</td>'
                    html += f'<td{rowspan} style="vertical-align:middle;">{escape(key[1])}</td>'
                    html += f'<td{rowspan} style="vertical-align:middle;">{escape(key[2])}</td>'
                    html += f'<td{rowspan} style="vertical-align:middle;">{escape(d.get("unit_name", "-"))}</td>'
                    html += f'<td{rowspan} style="vertical-align:middle;">¥{d.get("library_price", 0):.2f}</td>'
                html += f'<td>{escape(d.get("supplier_name", "-"))}</td>'
                html += f'<td>¥{d.get("this_price", 0):.2f}</td>'
                diff_color = '#e74c3c' if d.get('price_diff', 0) < 0 else '#27ae60'
                html += f'<td style="color:{diff_color}">¥{d.get("price_diff", 0):.2f}</td>'
                html += "</tr>\n"
        html += """                </tbody>
            </table>
"""

    # 计算每家供应商的拟定总额
    if items:
        supplier_totals = {}
        for item in items:
            quantity = float(item.get('quantity', 1) or 1)
            for quote in item.get('quotes', []):
                if quote.get('is_selected') == 1:
                    supplier_name = quote.get('supplier_name', '未知供应商')
                    tax_price = float(quote.get('tax_price', 0) or 0)
                    supplier_totals[supplier_name] = supplier_totals.get(supplier_name, 0) + tax_price * quantity
    else:
        supplier_totals = {}

    supplier_totals_html = ''
    if supplier_totals:
        supplier_totals_html = '<div style="flex:1;text-align:left;">' + \
            ' | '.join(f'{escape(name)}: <strong>¥{amt:,.2f}</strong>' for name, amt in supplier_totals.items()) + \
            '</div>'

    html += f"""<div class="total-section" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">
                {supplier_totals_html}
                <div style="text-align:right;">
                    <strong style="font-size:14px;">总金额：¥{inquiry.get('total_amount', 0):.2f}</strong>
                    <div class="amount-chinese">大写：{amount_chinese}</div>
                </div>
            </div>

            <div class="approval-section">
                <div class="approval-title">审批记录</div>
                <table class="approval-table">
                    <thead>
                        <tr>
                            <th style="width: 100px;">审批级别</th>
                            <th>审批人</th>
                            <th>结果</th>
                            <th>备注</th>
                            <th style="width: 150px;">时间</th>
                        </tr>
                    </thead>
                    <tbody>
"""

    approval_levels = [
        {'level': '项目管理处', 'matches': ['主管同意', '主管已审']},
    ]

    for level_info in approval_levels:
        found = next((r for r in approval_records if any(m in r.get('result', '') for m in level_info['matches'])), None)
        if found:
            html += f"""
                        <tr>
                            <td>{level_info['level']}</td>
                            <td>{escape(found.get('approver_name', '-'))}</td>
                            <td style="color: #27ae60;">{escape(found.get('result', '-'))}</td>
                            <td>{escape(found.get('remark') or '-')}</td>
                            <td>{escape(found.get('approval_time', '-'))}</td>
                        </tr>
"""
        else:
            html += f"""
                        <tr>
                            <td>{level_info['level']}</td>
                            <td colspan="4" style="color: #f39c12; font-style: italic;">待审批</td>
                        </tr>
"""

    html += """                    </tbody>
                </table>
            </div>

            <div style="margin-top: 10px; font-size: 12px; color: #666;">
                <strong>备注：</strong>""" + escape(inquiry.get('remark') or '无') + """
            </div>

            <div class="signatures" style="margin-top:20px;">
                <div class="signature-item">
                    <div class="signature-line" style="margin-top:20px;">申请人</div>
                </div>
                <div class="signature-item">
                    <div class="signature-line" style="margin-top:20px;">材料员签字</div>
                </div>
                <div class="signature-item">
                    <div class="signature-line" style="margin-top:20px;">项目经理（执行经理、生产经理）</div>
                </div>
                <div class="signature-item">
                    <div class="signature-line" style="margin-top:20px;">日期</div>
                </div>
            </div>

            <div style="margin-top: 10px; font-size: 10px; color: #999; text-align: right;">
                打印时间：""" + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """
            </div>
        </div>
    </body>
    </html>
    """

    response = make_response(html)
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/export-supplier-orders', methods=['GET'])
def export_supplier_orders(inquiry_id):
    """导出供应商供货单（按供应商分sheet的Excel）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO

    conn = get_db()
    cursor = conn.cursor()

    # 获取询价单信息
    cursor.execute("""
        SELECT pi.*, u.real_name as applicant_name, p.project_code, p.project_name
        FROM purchase_inquiries pi
        LEFT JOIN users u ON pi.applicant_id = u.id
        LEFT JOIN projects p ON pi.project_id = p.id
        WHERE pi.id = ?
    """, (inquiry_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    inquiry = dict(row)
    project_display = format_project_display(
        inquiry.get('project_code'),
        inquiry.get('project_name'),
    )

    # 获取所有items
    cursor.execute("""
        SELECT i.id as item_id, i.quantity, i.material_id,
               m.material_name, m.specification, m.material_code,
               COALESCE(i.detail_spec, m.detail_spec, '') AS detail_spec,
               COALESCE(i.brand, m.brand, '') AS brand,
               u.unit_name
        FROM purchase_inquiry_items i
        LEFT JOIN materials m ON m.id = i.material_id
        LEFT JOIN units u ON u.id = m.unit_id
        WHERE i.inquiry_id = ?
        ORDER BY i.id
    """, (inquiry_id,))
    all_items = [dict(row) for row in cursor.fetchall()]

    if not all_items:
        conn.close()
        return jsonify({'success': False, 'message': '没有询价材料明细'})

    # 为每个item查找报价：优先 is_selected=1，否则取 is_lowest=1，最后取任意有效报价
    items_with_quotes = []
    for item in all_items:
        item_id = item['item_id']

        # 优先查找 is_selected=1 的报价
        cursor.execute("""
            SELECT q.supplier_id, q.tax_price, q.tax_exempt_price, q.tax_rate,
                   s.supplier_name
            FROM purchase_inquiry_quotes q
            LEFT JOIN suppliers s ON s.id = q.supplier_id
            WHERE q.item_id = ? AND q.is_selected = 1
            LIMIT 1
        """, (item_id,))
        quote = cursor.fetchone()

        # 如果没有 is_selected，退而查找 is_lowest=1 的最低价报价
        if not quote:
            cursor.execute("""
                SELECT q.supplier_id, q.tax_price, q.tax_exempt_price, q.tax_rate,
                       s.supplier_name
                FROM purchase_inquiry_quotes q
                LEFT JOIN suppliers s ON s.id = q.supplier_id
                WHERE q.item_id = ? AND q.is_lowest = 1
                LIMIT 1
            """, (item_id,))
            quote = cursor.fetchone()

        # 最后兜底：取任意一条有效报价
        if not quote:
            cursor.execute("""
                SELECT q.supplier_id, q.tax_price, q.tax_exempt_price, q.tax_rate,
                       s.supplier_name
                FROM purchase_inquiry_quotes q
                LEFT JOIN suppliers s ON s.id = q.supplier_id
                WHERE q.item_id = ? AND q.tax_price > 0
                LIMIT 1
            """, (item_id,))
            quote = cursor.fetchone()

        if quote:
            quote = dict(quote)
            item['supplier_id'] = quote['supplier_id']
            item['tax_price'] = quote['tax_price']
            item['tax_exempt_price'] = quote.get('tax_exempt_price', 0)
            item['tax_rate'] = quote.get('tax_rate', 0.13)
            item['supplier_name'] = quote.get('supplier_name', '')
            items_with_quotes.append(item)

    conn.close()

    if not items_with_quotes:
        return jsonify({'success': False, 'message': '没有有效的供应商报价数据'})

    # 按供应商分组
    supplier_groups = {}
    for r in items_with_quotes:
        supplier_id = r.get('supplier_id')
        if not supplier_id:
            continue
        supplier_name = r.get('supplier_name', '未知供应商')
        if supplier_id not in supplier_groups:
            supplier_groups[supplier_id] = {
                'name': supplier_name,
                'items': []
            }
        supplier_groups[supplier_id]['items'].append(r)

    if not supplier_groups:
        return jsonify({'success': False, 'message': '没有有效的供应商数据'})

    # 创建Excel工作簿
    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 定义样式
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    for supplier_id, group in supplier_groups.items():
        # Sheet名称（最长31字符，不能包含特殊字符）
        sheet_name = group['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
        ws = wb.create_sheet(title=sheet_name)

        # 标题行
        ws.merge_cells('A1:J1')
        ws['A1'] = f'供货清单 - {group["name"]}'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # 询价单信息
        ws['A2'] = f'询价单号：{inquiry["inquiry_no"]}'
        ws['A3'] = f'日期：{inquiry.get("inquiry_date", "-")}'
        ws['A4'] = f'项目：{project_display}'

        # 表头
        headers = ['序号', '材料编码', '材料名称', '规格型号', '详细规格', '品牌', '单位', '数量', '含税单价', '金额']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # 数据行
        total_amount = 0
        for idx, item in enumerate(group['items'], 1):
            quantity = item.get('quantity', 0)
            tax_price = item.get('tax_price', 0)
            amount = quantity * tax_price
            total_amount += amount

            row_data = [
                idx,
                item.get('material_code', ''),
                item.get('material_name', ''),
                item.get('specification', ''),
                item.get('detail_spec', ''),
                item.get('brand', ''),
                item.get('unit_name', ''),
                quantity,
                tax_price,
                amount
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=6 + idx, column=col, value=value)
                cell.border = thin_border
                if col in (8, 9, 10):
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                elif col == 1:
                    cell.alignment = center_align

        # 汇总行
        total_row = 6 + len(group['items']) + 1
        ws.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
        total_cell = ws.cell(row=total_row, column=10, value=total_amount)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        total_cell.border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    from urllib.parse import quote
    filename = f'供货清单_{inquiry["inquiry_no"]}.xlsx'
    encoded_filename = quote(filename)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response


# ==================== 草稿（暂存）功能 ====================

@inquiry_bp.route('/purchase-inquiries/draft', methods=['POST'])
def save_draft():
    """保存询价单草稿（支持新建和覆盖已有草稿）"""
    import json as json_module
    import sqlite3

    try:
        data = json_module.loads(request.data)
    except Exception as e:
        return jsonify({'success': False, 'message': 'JSON解析失败: ' + str(e)})

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    # 使用独立连接，避免 Flask g.db 生命周期问题
    conn = sqlite3.connect(config.DATABASE_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.text_factory = str
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    cursor = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    draft_id = data.get('draft_id')  # 如果有 draft_id 则是覆盖已有草稿
    items_data = data.get('items', [])
    project_id = data.get('project_id')
    inquiry_date = data.get('inquiry_date', now[:10])
    remark = data.get('remark', '')

    try:
        if draft_id:
            # 覆盖已有草稿：先删除旧 items/quotes
            cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ? AND applicant_id = ?",
                           (draft_id, user['id']))
            existing = cursor.fetchone()
            if not existing:
                conn.close()
                return jsonify({'success': False, 'message': '草稿不存在或无权编辑'})
            if dict(existing)['approval_status'] != '草稿':
                conn.close()
                return jsonify({'success': False, 'message': '该询价单不是草稿状态'})

            cursor.execute("SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?", (draft_id,))
            old_item_ids = [r[0] for r in cursor.fetchall()]
            if old_item_ids:
                placeholders = ','.join('?' * len(old_item_ids))
                cursor.execute(f"DELETE FROM purchase_inquiry_quotes WHERE item_id IN ({placeholders})", old_item_ids)
            cursor.execute("DELETE FROM purchase_inquiry_items WHERE inquiry_id = ?", (draft_id,))

            # 更新主表
            cursor.execute("""
                UPDATE purchase_inquiries
                SET inquiry_date = ?, project_id = ?, remark = ?, create_time = ?
                WHERE id = ?
            """, (inquiry_date, project_id, remark, now, draft_id))

            inquiry_id = draft_id
            inquiry_no = dict(existing)['inquiry_no']
        else:
            # 新建草稿
            inquiry_no = None
            inquiry_id = None

            cursor.execute("PRAGMA table_info(purchase_inquiries)")
            columns = [row[1] for row in cursor.fetchall()]
            has_project_id = 'project_id' in columns

            for attempt in range(5):
                try:
                    if project_id:
                        inquiry_no = generate_inquiry_no_by_project(project_id, inquiry_date)
                    else:
                        inquiry_no = generate_inquiry_no()

                    if has_project_id:
                        cursor.execute("""
                            INSERT INTO purchase_inquiries (
                                inquiry_no, inquiry_date, applicant_id, project_id, total_amount,
                                is_below_library_price, approval_status, create_time, remark
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (inquiry_no, inquiry_date, user['id'], project_id, 0, 0, '草稿', now, remark))
                    else:
                        cursor.execute("""
                            INSERT INTO purchase_inquiries (
                                inquiry_no, inquiry_date, applicant_id, total_amount,
                                is_below_library_price, approval_status, create_time, remark
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (inquiry_no, inquiry_date, user['id'], 0, 0, '草稿', now, remark))
                    inquiry_id = cursor.lastrowid
                    break
                except Exception as e:
                    if 'UNIQUE constraint failed' in str(e) and attempt < 4:
                        conn.rollback()
                        continue
                    raise

        if inquiry_id is None:
            conn.close()
            return jsonify({'success': False, 'message': '无法生成唯一单号，请稍后重试'})

        # 写入 items + quotes（草稿允许无材料）
        for item in items_data:
            material_id = item.get('material_id')
            quantity = float(item.get('quantity', 1) or 1)
            library_price = float(item.get('library_price', 0) or 0)
            selected_quote_id = item.get('selected_quote_id')
            tax_rate = float(item.get('tax_rate', 0.01) or 0.01)
            is_national_standard = item.get('is_national_standard', 0)
            is_cash_price = item.get('is_cash_price', 0)
            detail_spec = item.get('detail_spec', '') or '常规'
            brand = item.get('brand', '') or '无'

            if is_cash_price:
                tax_rate = 0.01

            cursor.execute("""
                INSERT INTO purchase_inquiry_items (
                    inquiry_id, material_id, quantity, library_price,
                    selected_quote_id, tax_rate, is_national_standard, is_cash_price,
                    detail_spec, brand, create_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                inquiry_id,
                material_id if material_id else None,
                quantity, library_price, selected_quote_id,
                tax_rate, is_national_standard, is_cash_price,
                detail_spec, brand, now
            ))
            item_id = cursor.lastrowid

            quotes = item.get('quotes', [])
            for quote in quotes:
                supplier_id = quote.get('supplier_id')
                if not supplier_id:
                    continue
                tax_price = float(quote.get('tax_price', 0) or 0)
                tax_exempt_price = float(quote.get('tax_exempt_price', 0) or 0)
                tax_rate_val = float(quote.get('tax_rate', 0.13) or 0.13)
                total = tax_price * quantity

                if tax_exempt_price == 0 and tax_price > 0:
                    tax_exempt_price = round(tax_price / (1 + tax_rate_val), 2)

                cursor.execute("""
                    INSERT INTO purchase_inquiry_quotes (
                        item_id, supplier_id, tax_price, tax_exempt_price,
                        tax_rate, total_amount, is_lowest, is_selected, create_time
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item_id, supplier_id, tax_price, tax_exempt_price,
                    tax_rate_val, total, 0,
                    1 if selected_quote_id and str(quote.get('supplier_id')) == str(selected_quote_id) else 0,
                    now
                ))

        conn.commit()
        return jsonify({
            'success': True,
            'message': '草稿已保存',
            'draft_id': inquiry_id,
            'inquiry_no': inquiry_no
        })

    except Exception as e:
        logger.error("保存草稿失败: %s", e)
        import traceback
        traceback.print_exc()
        conn.rollback()
        return jsonify({'success': False, 'message': '保存草稿失败: ' + str(e)})


@inquiry_bp.route('/purchase-inquiries/drafts', methods=['GET'])
def get_drafts():
    """获取当前用户的草稿列表"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pi.*, u.real_name as applicant_name
        FROM purchase_inquiries pi
        LEFT JOIN users u ON pi.applicant_id = u.id
        WHERE pi.applicant_id = ? AND pi.approval_status = '草稿'
        ORDER BY pi.create_time DESC
    """, (user['id'],))
    drafts = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': drafts})


@inquiry_bp.route('/purchase-inquiries/draft/<int:draft_id>/export-quote-sheet', methods=['GET'])
def export_draft_quote_sheet(draft_id):
    """导出草稿询价表，供材料员发给供应商填写报价。"""
    from io import BytesIO
    from urllib.parse import quote as url_quote
    from zipfile import ZIP_DEFLATED, ZipFile
    from xml.sax.saxutils import escape as xml_escape

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pi.*, p.project_code, p.project_name
        FROM purchase_inquiries pi
        LEFT JOIN projects p ON pi.project_id = p.id
        WHERE pi.id = ?
    """, (draft_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '草稿不存在'}), 404

    draft = dict(row)
    if draft.get('approval_status') != '草稿':
        conn.close()
        return jsonify({'success': False, 'message': '该询价单不是草稿'})
    if draft.get('applicant_id') != user.get('id'):
        conn.close()
        return jsonify({'success': False, 'message': '只有申请人可以导出此草稿'}), 403

    def table_columns(table_name):
        cursor.execute(f"PRAGMA table_info({table_name})")
        return {col[1] for col in cursor.fetchall()}

    item_columns = table_columns('purchase_inquiry_items')
    material_columns = table_columns('materials')
    item_detail_expr = 'i.detail_spec' if 'detail_spec' in item_columns else 'NULL'
    material_detail_expr = 'm.detail_spec' if 'detail_spec' in material_columns else 'NULL'
    item_brand_expr = 'i.brand' if 'brand' in item_columns else 'NULL'
    material_brand_expr = 'm.brand' if 'brand' in material_columns else 'NULL'
    if 'is_national_standard' in item_columns:
        national_standard_expr = 'i.is_national_standard'
    elif 'is_national_standard' in material_columns:
        national_standard_expr = 'm.is_national_standard'
    else:
        national_standard_expr = '0'

    cursor.execute(f"""
        SELECT i.id AS item_id, i.quantity, {national_standard_expr} AS is_national_standard,
               COALESCE(m.specification, '') AS export_spec,
               COALESCE({item_detail_expr}, {material_detail_expr}, '') AS export_detail_spec,
               COALESCE({item_brand_expr}, {material_brand_expr}, '') AS export_brand,
               m.material_name, m.specification, u.unit_name
        FROM purchase_inquiry_items i
        LEFT JOIN materials m ON m.id = i.material_id
        LEFT JOIN units u ON u.id = m.unit_id
        WHERE i.inquiry_id = ?
        ORDER BY i.id
    """, (draft_id,))
    items = [dict(r) for r in cursor.fetchall()]
    if not items:
        conn.close()
        return jsonify({'success': False, 'message': '草稿没有询价材料明细'})

    item_ids = [item['item_id'] for item in items]
    placeholders = ','.join('?' * len(item_ids))
    cursor.execute(f"""
        SELECT q.item_id, q.supplier_id, q.tax_rate, s.supplier_name
        FROM purchase_inquiry_quotes q
        LEFT JOIN suppliers s ON s.id = q.supplier_id
        WHERE q.item_id IN ({placeholders}) AND q.supplier_id IS NOT NULL
        ORDER BY q.id
    """, item_ids)
    suppliers = []
    seen_suppliers = set()
    for quote_row in cursor.fetchall():
        quote_data = dict(quote_row)
        supplier_id = quote_data.get('supplier_id')
        if supplier_id in seen_suppliers:
            continue
        seen_suppliers.add(supplier_id)
        suppliers.append({
            'id': supplier_id,
            'name': quote_data.get('supplier_name') or '供应商',
            'tax_rate': quote_data.get('tax_rate') if quote_data.get('tax_rate') is not None else 0.01,
        })
    conn.close()

    if not suppliers:
        suppliers = [{'id': None, 'name': '供应商', 'tax_rate': 0.01}]

    def tax_label(rate):
        try:
            return f"{int(round(float(rate) * 100))}%专票"
        except (TypeError, ValueError):
            return '专票'

    base_cols = 8
    total_cols = base_cols + len(suppliers) * 2

    def col_name(index):
        name = ''
        while index:
            index, rem = divmod(index - 1, 26)
            name = chr(65 + rem) + name
        return name

    def cell_ref(row, col):
        return f'{col_name(col)}{row}'

    def cell_xml(row, col, value=None, style=1, formula=None):
        ref = cell_ref(row, col)
        style_attr = f' s="{style}"' if style else ''
        if formula:
            return f'<c r="{ref}"{style_attr}><f>{xml_escape(formula)}</f></c>'
        if value is None or value == '':
            return f'<c r="{ref}"{style_attr}/>'
        if isinstance(value, (int, float)):
            return f'<c r="{ref}"{style_attr}><v>{value}</v></c>'
        text = xml_escape(str(value))
        return f'<c r="{ref}" t="inlineStr"{style_attr}><is><t xml:space="preserve">{text}</t></is></c>'

    last_col = col_name(total_cols)
    project_display = format_project_display(draft.get('project_code'), draft.get('project_name'))
    headers = ['序号', '材料名称', '规格型号', '详细规格', '品牌', '是否国标', '单位', '数量']
    for supplier in suppliers:
        headers.extend([f'{supplier["name"]}单价{tax_label(supplier.get("tax_rate"))}', f'{supplier["name"]}总价'])

    rows_xml = []
    rows_xml.append(
        '<row r="1" ht="46" customHeight="1">'
        + cell_xml(1, 1, '零星材采购比价表', style=2)
        + '</row>'
    )
    rows_xml.append(
        '<row r="2" ht="22" customHeight="1">'
        + cell_xml(2, 1, f'项目名称：{project_display}', style=1)
        + cell_xml(2, base_cols + 1, f'时间：{draft.get("inquiry_date") or ""}', style=1)
        + '</row>'
    )
    rows_xml.append(
        '<row r="3" ht="42" customHeight="1">'
        + ''.join(cell_xml(3, col, header, style=3) for col, header in enumerate(headers, 1))
        + '</row>'
    )

    data_rows = []
    for idx, item in enumerate(items, 1):
        row_idx = idx + 3
        row_values = [
            idx,
            item.get('material_name') or '',
            item.get('export_spec') or item.get('specification') or '',
            item.get('export_detail_spec') or '',
            item.get('export_brand') or '',
            '是' if item.get('is_national_standard') else '否',
            item.get('unit_name') or '',
            item.get('quantity') or 0,
        ]
        cells = [cell_xml(row_idx, col, value, style=1) for col, value in enumerate(row_values, 1)]
        for supplier_idx, _supplier in enumerate(suppliers):
            price_col = base_cols + 1 + supplier_idx * 2
            total_col = price_col + 1
            price_ref = cell_ref(row_idx, price_col)
            cells.append(cell_xml(row_idx, price_col, None, style=4))
            cells.append(cell_xml(row_idx, total_col, style=4, formula=f'IF({price_ref}="","",{price_ref}*$H{row_idx})'))
        data_rows.append(f'<row r="{row_idx}" ht="25" customHeight="1">{"".join(cells)}</row>')
    rows_xml.extend(data_rows)

    col_widths = {1: 4, 2: 24, 3: 20, 4: 28, 5: 10, 6: 10, 7: 10, 8: 9}
    cols_xml = ''.join(
        f'<col min="{idx}" max="{idx}" width="{col_widths.get(idx, 16 if idx % 2 else 12)}" customWidth="1"/>'
        for idx in range(1, total_cols + 1)
    )
    merges_xml = (
        '<mergeCells count="3">'
        f'<mergeCell ref="A1:{last_col}1"/>'
        f'<mergeCell ref="A2:{col_name(base_cols)}2"/>'
        f'<mergeCell ref="{col_name(base_cols + 1)}2:{last_col}2"/>'
        '</mergeCells>'
    )
    sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="3" topLeftCell="A4" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <cols>{cols_xml}</cols>
  <sheetData>{''.join(rows_xml)}</sheetData>
  {merges_xml}
  <autoFilter ref="A3:{last_col}{3 + len(items)}"/>
</worksheet>'''

    styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3"><font><sz val="11"/><name val="宋体"/></font><font><b/><sz val="18"/><name val="宋体"/></font><font><b/><sz val="11"/><name val="宋体"/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFD9EAF7"/><bgColor indexed="64"/></patternFill></fill></fills>
  <borders count="2"><border/><border><left style="thin"/><right style="thin"/><top style="thin"/><bottom style="thin"/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="5">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="1" fillId="0" borderId="1" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="4" fontId="0" fillId="0" borderId="1" xfId="0" applyNumberFormat="1" applyBorder="1" applyAlignment="1"><alignment horizontal="right" vertical="center"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''

    output = BytesIO()
    with ZipFile(output, 'w', ZIP_DEFLATED) as xlsx:
        xlsx.writestr('[Content_Types].xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>''')
        xlsx.writestr('_rels/.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>''')
        xlsx.writestr('xl/workbook.xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="询价表" sheetId="1" r:id="rId1"/></sheets>
</workbook>''')
        xlsx.writestr('xl/_rels/workbook.xml.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>''')
        xlsx.writestr('xl/worksheets/sheet1.xml', sheet_xml)
        xlsx.writestr('xl/styles.xml', styles_xml)
    output.seek(0)

    filename = f'询价表_{draft.get("inquiry_no") or draft_id}.xlsx'
    encoded_filename = url_quote(filename)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response


@inquiry_bp.route('/purchase-inquiries/draft/<int:draft_id>/import-quote-sheet', methods=['POST'])
def import_draft_quote_sheet(draft_id):
    """导入草稿询比价表，解析为前端询价明细数据。"""
    from io import BytesIO
    import re

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    upload = request.files.get('file')
    if not upload:
        return jsonify({'success': False, 'message': '请选择要导入的询比价表'})

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return jsonify({'success': False, 'message': '服务器缺少 Excel 解析依赖 openpyxl'})

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (draft_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '草稿不存在'}), 404

    draft = dict(row)
    if draft.get('approval_status') != '草稿':
        conn.close()
        return jsonify({'success': False, 'message': '该询价单不是草稿状态'})
    if draft.get('applicant_id') != user.get('id'):
        conn.close()
        return jsonify({'success': False, 'message': '只有申请人可以导入此草稿'}), 403

    try:
        workbook = load_workbook(BytesIO(upload.read()), data_only=True)
        sheet = workbook.active
    except Exception:
        conn.close()
        return jsonify({'success': False, 'message': '无法读取 Excel，请使用询比价导出的模板导入'})

    headers = [sheet.cell(3, col).value for col in range(1, sheet.max_column + 1)]
    required_headers = ['材料名称', '规格型号', '详细规格', '品牌', '单位', '数量']
    if not all(header in headers for header in required_headers):
        conn.close()
        return jsonify({'success': False, 'message': '模板表头不正确，请使用询比价导出的模板导入'})

    def normalize(value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def find_supplier(name):
        """精确匹配 → 去括号匹配 → LIKE 模糊匹配"""
        cursor.execute("SELECT id, supplier_name FROM suppliers WHERE supplier_name = ? LIMIT 1", (name,))
        row = cursor.fetchone()
        if row:
            return dict(row)
        stripped = name.rstrip('（(').rstrip()
        if stripped and stripped != name:
            cursor.execute("SELECT id, supplier_name FROM suppliers WHERE supplier_name = ? LIMIT 1", (stripped,))
            row = cursor.fetchone()
            if row:
                return dict(row)
        cursor.execute("SELECT id, supplier_name FROM suppliers WHERE supplier_name LIKE ? LIMIT 1", (f'%{name}%',))
        row = cursor.fetchone()
        if row:
            return dict(row)
        if stripped and stripped != name:
            cursor.execute("SELECT id, supplier_name FROM suppliers WHERE supplier_name LIKE ? LIMIT 1", (f'%{stripped}%',))
            row = cursor.fetchone()
            if row:
                return dict(row)
        return None

    def to_number(value, default=0):
        if value is None or value == '':
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def parse_yes_no(value):
        text = normalize(value)
        return 1 if text in ('是', '1', 'true', 'TRUE', 'True') else 0

    def parse_supplier_header(value):
        text = normalize(value)
        if '单价' not in text:
            return None
        match = re.match(r'^(.*?)单价(?:(\d+(?:\.\d+)?)%)?', text)
        if not match:
            return None
        name = match.group(1).strip().rstrip('（(')
        if not name:
            return None
        rate_text = match.group(2)
        rate = float(rate_text) / 100 if rate_text else 0.13
        return {'name': name, 'tax_rate': rate}

    supplier_columns = []
    for col in range(9, sheet.max_column + 1, 2):
        supplier_info = parse_supplier_header(sheet.cell(3, col).value)
        if supplier_info:
            supplier_info['price_col'] = col
            supplier_columns.append(supplier_info)

    warnings = []
    parsed_items = []
    for row_idx in range(4, sheet.max_row + 1):
        material_name = normalize(sheet.cell(row_idx, 2).value)
        specification = normalize(sheet.cell(row_idx, 3).value)
        detail_spec = normalize(sheet.cell(row_idx, 4).value)
        brand = normalize(sheet.cell(row_idx, 5).value)
        is_national_standard = parse_yes_no(sheet.cell(row_idx, 6).value)
        unit_name = normalize(sheet.cell(row_idx, 7).value)
        quantity = to_number(sheet.cell(row_idx, 8).value, 1)

        if not any([material_name, specification, detail_spec, brand, unit_name]):
            continue

        # 空值填充默认值
        detail_spec = detail_spec or '常规'
        brand = brand or '无'

        # 归一化规格：去掉末尾 .0，去掉 DN 前缀
        def normalize_spec(s):
            s = s.strip()
            if s.endswith('.0'):
                s = s[:-2]
            if s.upper().startswith('DN'):
                s = s[2:]
            return s

        spec_norm = normalize_spec(specification)

        # 第一轮：精确匹配（名称 + 规格 + 单位）
        cursor.execute("""
            SELECT m.id, m.material_code, m.material_name, m.specification,
                   m.detail_spec, m.brand, m.tax_price, m.cash_price,
                   COALESCE(m.is_cash_price, 0) AS is_cash_price,
                   u.unit_name
            FROM materials m
            LEFT JOIN units u ON u.id = m.unit_id
            WHERE m.material_name = ?
              AND COALESCE(m.specification, '') = ?
              AND COALESCE(u.unit_name, '') = ?
            LIMIT 1
        """, (material_name, specification, unit_name))
        material_row = cursor.fetchone()
        material = dict(material_row) if material_row else None

        # 第二轮：归一化规格匹配（解决 "25" vs "25.0" vs "DN25"）
        if not material:
            cursor.execute("""
                SELECT m.id, m.material_code, m.material_name, m.specification,
                       m.detail_spec, m.brand, m.tax_price, m.cash_price,
                       COALESCE(m.is_cash_price, 0) AS is_cash_price,
                       u.unit_name
                FROM materials m
                LEFT JOIN units u ON u.id = m.unit_id
                WHERE m.material_name = ?
                  AND COALESCE(u.unit_name, '') = ?
            """, (material_name, unit_name))
            for row in cursor.fetchall():
                if normalize_spec(row['specification'] or '') == spec_norm:
                    material = dict(row)
                    break

        # 第三轮：名称 + 单位匹配（忽略规格）
        if not material:
            cursor.execute("""
                SELECT m.id, m.material_code, m.material_name, m.specification,
                       m.detail_spec, m.brand, m.tax_price, m.cash_price,
                       COALESCE(m.is_cash_price, 0) AS is_cash_price,
                       u.unit_name
                FROM materials m
                LEFT JOIN units u ON u.id = m.unit_id
                WHERE m.material_name = ?
                  AND COALESCE(u.unit_name, '') = ?
                LIMIT 1
            """, (material_name, unit_name))
            material_row = cursor.fetchone()
            material = dict(material_row) if material_row else None

        # 第四轮：仅按名称匹配（忽略单位和规格）
        if not material:
            cursor.execute("""
                SELECT m.id, m.material_code, m.material_name, m.specification,
                       m.detail_spec, m.brand, m.tax_price, m.cash_price,
                       COALESCE(m.is_cash_price, 0) AS is_cash_price,
                       u.unit_name
                FROM materials m
                LEFT JOIN units u ON u.id = m.unit_id
                WHERE m.material_name = ?
                LIMIT 1
            """, (material_name,))
            material_row = cursor.fetchone()
            material = dict(material_row) if material_row else None

        # 第五轮：模糊名称匹配（名称前2字 + 单位）
        if not material and len(material_name) >= 2:
            cursor.execute("""
                SELECT m.id, m.material_code, m.material_name, m.specification,
                       m.detail_spec, m.brand, m.tax_price, m.cash_price,
                       COALESCE(m.is_cash_price, 0) AS is_cash_price,
                       u.unit_name
                FROM materials m
                LEFT JOIN units u ON u.id = m.unit_id
                WHERE m.material_name LIKE ?
                  AND COALESCE(u.unit_name, '') = ?
                LIMIT 1
            """, (f'%{material_name[:2]}%', unit_name))
            material_row = cursor.fetchone()
            material = dict(material_row) if material_row else None

        # 第六轮：模糊名称匹配（忽略单位）
        if not material and len(material_name) >= 2:
            cursor.execute("""
                SELECT m.id, m.material_code, m.material_name, m.specification,
                       m.detail_spec, m.brand, m.tax_price, m.cash_price,
                       COALESCE(m.is_cash_price, 0) AS is_cash_price,
                       u.unit_name
                FROM materials m
                LEFT JOIN units u ON u.id = m.unit_id
                WHERE m.material_name LIKE ?
                LIMIT 1
            """, (f'%{material_name[:2]}%',))
            material_row = cursor.fetchone()
            material = dict(material_row) if material_row else None

        # 所有匹配均失败，自动创建新材料
        item_warnings = []
        if not material:
            cursor.execute("SELECT id FROM units WHERE unit_name = ? LIMIT 1", (unit_name,))
            unit_row = cursor.fetchone()
            unit_id = unit_row[0] if unit_row else None

            # 生成材料编码
            cursor.execute("SELECT material_code FROM materials ORDER BY id DESC LIMIT 1")
            last_row = cursor.fetchone()
            if last_row and last_row[0]:
                try:
                    parts = last_row[0].rsplit('-', 1)
                    new_num = int(parts[1]) + 1
                    new_code = f"{parts[0]}-{new_num:03d}"
                except (ValueError, IndexError):
                    new_code = f"LX-{(last_row[0] or '0').replace('LX-','').replace('-','')}"
            else:
                new_code = "LX-001"

            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                INSERT INTO materials (
                    material_code, material_name, specification, detail_spec, brand,
                    unit_id, tax_price, is_cash_price, cash_price, create_time, tax_rate
                ) VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, ?, 0.01)
            """, (new_code, material_name, specification, detail_spec, brand, unit_id, now_str))
            new_mat_id = cursor.lastrowid

            material = {
                'id': new_mat_id,
                'material_code': new_code,
                'material_name': material_name,
                'specification': specification,
                'detail_spec': detail_spec,
                'brand': brand,
                'tax_price': 0,
                'cash_price': 0,
                'is_cash_price': 0,
                'unit_name': unit_name,
            }
            item_warnings.append(f'第{row_idx}行材料已自动创建：{material_name}')

        quotes = []
        for supplier in supplier_columns:
            supplier_name = supplier['name']
            supplier_data = find_supplier(supplier_name)
            tax_price = to_number(sheet.cell(row_idx, supplier['price_col']).value, 0)
            if not supplier_data:
                message = f'第{row_idx}行供应商未匹配：{supplier_name}'
                item_warnings.append(message)
                warnings.append(message)
                continue
            tax_rate = supplier['tax_rate']
            tax_exempt_price = round(tax_price / (1 + tax_rate), 2) if tax_price > 0 and tax_rate > -1 else 0
            quotes.append({
                'supplier_id': supplier_data['id'],
                'supplier_name': supplier_data['supplier_name'],
                'tax_price': round(tax_price, 2),
                'tax_exempt_price': tax_exempt_price,
                'tax_rate': tax_rate,
                'total_amount': round(tax_price * quantity, 2) if tax_price > 0 else 0,
                'is_lowest': 0,
                'is_selected': 0,
            })

        parsed_items.append({
            'material_id': material.get('id') if material else '',
            'material_code': material.get('material_code') if material else '',
            'material_name': material.get('material_name') if material else material_name,
            'specification': material.get('specification') if material else specification,
            'detail_spec': (material.get('detail_spec') or '常规') if material else detail_spec,
            'brand': (material.get('brand') or '无') if material else brand,
            'unit_name': material.get('unit_name') if material else unit_name,
            'quantity': quantity,
            'library_price': material.get('tax_price', 0) if material else 0,
            'tax_price': material.get('tax_price', 0) if material else 0,
            'cash_price': material.get('cash_price', 0) if material else 0,
            'is_cash_price': material.get('is_cash_price', 0) if material else 0,
            'is_national_standard': is_national_standard,
            'unmatched_material': material is None,
            'quotes': quotes,
            'warnings': item_warnings,
        })

    conn.close()

    if not parsed_items:
        return jsonify({'success': False, 'message': '未读取到有效的询价明细'})

    for item in parsed_items:
        for message in item.get('warnings', []):
            if message not in warnings:
                warnings.append(message)

    return jsonify({'success': True, 'items': parsed_items, 'warnings': warnings})


@inquiry_bp.route('/purchase-inquiries/draft/<int:draft_id>/submit', methods=['POST'])
def submit_draft(draft_id):
    """提交草稿（转为待审批）"""
    import json as json_module

    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    try:
        data = json_module.loads(request.data)
    except Exception as e:
        return jsonify({'success': False, 'message': 'JSON解析失败: ' + str(e)})

    conn = get_db()
    cursor = conn.cursor()

    # 校验草稿
    cursor.execute("SELECT * FROM purchase_inquiries WHERE id = ?", (draft_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '草稿不存在'})

    draft = dict(row)
    if draft['approval_status'] != '草稿':
        conn.close()
        return jsonify({'success': False, 'message': '该询价单不是草稿状态'})
    if draft['applicant_id'] != user['id']:
        conn.close()
        return jsonify({'success': False, 'message': '只有申请人可以提交此草稿'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 复用 create_inquiry 的校验逻辑
    items_data = data.get('items', [])
    if not items_data:
        conn.close()
        return jsonify({'success': False, 'message': '请添加询价材料'})

    try:
        # 删除旧 items 和 quotes
        cursor.execute("SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?", (draft_id,))
        old_item_ids = [r[0] for r in cursor.fetchall()]
        if old_item_ids:
            placeholders = ','.join('?' * len(old_item_ids))
            cursor.execute(f"DELETE FROM purchase_inquiry_quotes WHERE item_id IN ({placeholders})", old_item_ids)
        cursor.execute("DELETE FROM purchase_inquiry_items WHERE inquiry_id = ?", (draft_id,))

        # 计算总金额和是否低于库内价
        total_amount = 0
        is_below = 0
        for item in items_data:
            selected_id = item.get('selected_quote_id')
            quantity = float(item.get('quantity', 1) or 1)
            for quote in item.get('quotes', []):
                tax_price = float(quote.get('tax_price', 0) or 0)
                if selected_id and str(quote.get('supplier_id')) == str(selected_id):
                    total_amount += tax_price * quantity
                elif not selected_id and quote.get('is_lowest'):
                    total_amount += tax_price * quantity
            library_price = float(item.get('library_price', 0) or 0)
            for quote in item.get('quotes', []):
                tp = float(quote.get('tax_price', 0) or 0)
                if library_price > 0 and tp < library_price:
                    is_below = 1
                    break

        # 更新主表
        project_id = data.get('project_id', draft.get('project_id'))
        inquiry_date = data.get('inquiry_date', draft.get('inquiry_date', now[:10]))
        cursor.execute("""
            UPDATE purchase_inquiries
            SET inquiry_date = ?, project_id = ?, total_amount = ?,
                is_below_library_price = ?, approval_status = '待审批',
                remark = ?, create_time = ?
            WHERE id = ?
        """, (inquiry_date, project_id, total_amount, is_below,
              data.get('remark', draft.get('remark', '')), now, draft_id))

        # 写入新 items + quotes
        for item in items_data:
            material_id = item.get('material_id')
            quantity = float(item.get('quantity', 1) or 1)
            library_price = float(item.get('library_price', 0) or 0)
            selected_quote_id = item.get('selected_quote_id')
            tax_rate = float(item.get('tax_rate', 0.01) or 0.01)
            is_national_standard = item.get('is_national_standard', 0)
            is_cash_price = item.get('is_cash_price', 0)
            detail_spec = item.get('detail_spec', '') or '常规'
            brand = item.get('brand', '') or '无'

            if is_cash_price:
                tax_rate = 0.01

            cursor.execute("""
                INSERT INTO purchase_inquiry_items (
                    inquiry_id, material_id, quantity, library_price,
                    selected_quote_id, tax_rate, is_national_standard, is_cash_price,
                    detail_spec, brand, create_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                draft_id, material_id if material_id else None,
                quantity, library_price, selected_quote_id,
                tax_rate, is_national_standard, is_cash_price,
                detail_spec, brand, now
            ))
            item_id = cursor.lastrowid

            quotes = item.get('quotes', [])
            valid_quotes = [q for q in quotes if q.get('supplier_id')]

            if valid_quotes:
                priced_quotes = [(float(q.get('tax_exempt_price', 0) or float(q.get('tax_price', 0)) / (1 + float(q.get('tax_rate', 0.13) or 0.13))), i) for i, q in enumerate(valid_quotes) if float(q.get('tax_price', 0) or 0) > 0]
                lowest_idx = min(priced_quotes, key=lambda x: x[0])[1] if priced_quotes else -1

                for i, quote in enumerate(valid_quotes):
                    tp = float(quote.get('tax_price', 0) or 0)
                    tep = float(quote.get('tax_exempt_price', 0) or 0)
                    tr = float(quote.get('tax_rate', 0.13) or 0.13)
                    total = tp * quantity

                    if tep == 0 and tp > 0:
                        tep = round(tp / (1 + tr), 2)

                    q_status = 'submitted' if tp > 0 else 'pending'

                    cursor.execute("""
                        INSERT INTO purchase_inquiry_quotes (
                            item_id, supplier_id, tax_price, tax_exempt_price,
                            tax_rate, total_amount, is_lowest, is_selected, quote_status, create_time
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        item_id, quote.get('supplier_id'),
                        tp, tep, tr, total,
                        1 if i == lowest_idx else 0,
                        1 if selected_quote_id and str(quote.get('supplier_id')) == str(selected_quote_id) else 0,
                        q_status, now
                    ))

        # 记录提交审批
        cursor.execute("""
            INSERT INTO approval_records (order_type, order_id, approver_id, approver_name, result, remark, approval_time)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ('purchase_inquiry', draft_id, user['id'], user['real_name'], '提交审批', '', now))

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'inquiry_no': draft['inquiry_no'], 'id': draft_id})

    except Exception as e:
        logger.error("提交草稿失败: %s", e)
        import traceback
        traceback.print_exc()
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': '提交草稿失败: ' + str(e)})


# ==================== 供应商报价邀请 ====================

@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/publish-quotes', methods=['POST'])
def publish_quotes(inquiry_id):
    """发布报价邀请：为询比价单的材料+供应商组合创建待报价记录"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    # 检查询价单存在
    cursor.execute("SELECT id, quote_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    inquiry = cursor.fetchone()
    if not inquiry:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 获取该询价单的所有材料项
    cursor.execute("SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
    items = cursor.fetchall()
    if not items:
        conn.close()
        return jsonify({'success': False, 'message': '该询价单没有材料项'})

    created = 0
    for item_row in items:
        item_id = item_row['id']

        # 获取该材料项已有的供应商报价行
        cursor.execute("""
            SELECT supplier_id FROM purchase_inquiry_quotes WHERE item_id = ?
        """, (item_id,))
        existing_suppliers = {row['supplier_id'] for row in cursor.fetchall()}

        # 如果没有供应商信息，跳过（需要前端传入供应商列表）
        # 这里从已有报价行获取供应商，或从请求数据获取
        if not existing_suppliers:
            continue

        # 为已有供应商创建/更新待报价记录
        for supplier_id in existing_suppliers:
            cursor.execute("""
                SELECT id FROM purchase_inquiry_quotes
                WHERE item_id = ? AND supplier_id = ?
            """, (item_id, supplier_id))
            existing = cursor.fetchone()
            if existing:
                # 已存在，更新状态为 pending（如果还是初始状态）
                cursor.execute("""
                    UPDATE purchase_inquiry_quotes
                    SET quote_status = CASE WHEN quote_status IN ('locked') THEN quote_status ELSE 'pending' END,
                        updated_at = ?
                    WHERE id = ?
                """, (now, existing['id']))
            else:
                # 创建新的待报价记录
                cursor.execute("""
                    INSERT INTO purchase_inquiry_quotes (
                        item_id, supplier_id, tax_price, tax_exempt_price,
                        tax_rate, total_amount, is_lowest, is_selected,
                        quote_status, create_time
                    ) VALUES (?, ?, 0, 0, 0.13, 0, 0, 0, 'pending', ?)
                """, (item_id, supplier_id, now))
                created += 1

    # 更新询价单报价状态
    data = request.get_json(silent=True) or {}
    deadline = data.get('quote_deadline')
    cursor.execute("""
        UPDATE purchase_inquiries
        SET quote_status = 'collecting', quote_deadline = ?
        WHERE id = ?
    """, (deadline, inquiry_id))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': f'报价邀请已发布，新增 {created} 条待报价记录'})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/publish-quotes-to', methods=['POST'])
def publish_quotes_to_suppliers(inquiry_id):
    """发布报价邀请给指定供应商列表"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    data = request.get_json(silent=True) or {}
    supplier_ids = data.get('supplier_ids', [])
    if not supplier_ids:
        return jsonify({'success': False, 'message': '请选择至少一个供应商'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, quote_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    inquiry = cursor.fetchone()
    if not inquiry:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute("SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?", (inquiry_id,))
    items = cursor.fetchall()
    if not items:
        conn.close()
        return jsonify({'success': False, 'message': '该询价单没有材料项'})

    created = 0
    for item_row in items:
        item_id = item_row['id']
        for supplier_id in supplier_ids:
            cursor.execute("""
                SELECT id, quote_status FROM purchase_inquiry_quotes
                WHERE item_id = ? AND supplier_id = ?
            """, (item_id, supplier_id))
            existing = cursor.fetchone()
            if existing:
                if existing['quote_status'] == 'locked':
                    continue
                cursor.execute("""
                    UPDATE purchase_inquiry_quotes SET quote_status = 'pending', updated_at = ?
                    WHERE id = ?
                """, (now, existing['id']))
            else:
                cursor.execute("""
                    INSERT INTO purchase_inquiry_quotes (
                        item_id, supplier_id, tax_price, tax_exempt_price,
                        tax_rate, total_amount, is_lowest, is_selected,
                        quote_status, create_time
                    ) VALUES (?, ?, 0, 0, 0.13, 0, 0, 0, 'pending', ?)
                """, (item_id, supplier_id, now))
                created += 1

    deadline = data.get('quote_deadline')
    cursor.execute("""
        UPDATE purchase_inquiries
        SET quote_status = 'collecting', quote_deadline = ?
        WHERE id = ?
    """, (deadline, inquiry_id))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': f'已发布给 {len(supplier_ids)} 个供应商，新增 {created} 条报价记录'})


@inquiry_bp.route('/purchase-inquiries/<int:inquiry_id>/lock-quotes', methods=['POST'])
def lock_quotes(inquiry_id):
    """锁定报价：供应商不可再修改"""
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '未登录'})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, quote_status FROM purchase_inquiries WHERE id = ?", (inquiry_id,))
    inquiry = cursor.fetchone()
    if not inquiry:
        conn.close()
        return jsonify({'success': False, 'message': '询价单不存在'})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 锁定所有报价行
    cursor.execute("""
        UPDATE purchase_inquiry_quotes
        SET quote_status = 'locked', updated_at = ?
        WHERE item_id IN (SELECT id FROM purchase_inquiry_items WHERE inquiry_id = ?)
          AND quote_status != 'locked'
    """, (now, inquiry_id))

    # 更新询价单报价状态
    cursor.execute("""
        UPDATE purchase_inquiries SET quote_status = 'locked' WHERE id = ?
    """, (inquiry_id,))

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '报价已锁定'})
